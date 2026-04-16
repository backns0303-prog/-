import calendar
import io
import json
import os
import math
from collections.abc import Mapping
from datetime import date, datetime, timedelta
from functools import lru_cache
from pathlib import Path
import re
from time import perf_counter
from zoneinfo import ZoneInfo

import gspread
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from gspread.exceptions import WorksheetNotFound
from google.oauth2.service_account import Credentials


st.set_page_config(page_title="품평/프로젝트 관리", layout="wide")


COMPONENT_DIR = Path(__file__).parent / "components" / "overlay_calendar_component"
overlay_calendar_component = components.declare_component(
    "overlay_calendar_component",
    path=str(COMPONENT_DIR),
)
SUMMARY_COMPONENT_DIR = Path(__file__).parent / "components" / "summary_cards_component"
summary_cards_component = components.declare_component(
    "summary_cards_component",
    path=str(SUMMARY_COMPONENT_DIR),
)

SHOW_PERF_LOG_PANEL = False

GOOGLE_CREDENTIALS_FILE = Path(__file__).parent / "streamlit-sheets-upload-34b193fd0a59.json"
GOOGLE_SPREADSHEET_ID = "1Jy1DFHveJYFEw2lVg_pUGeE7HCcFmYaeUb6FwSrZGJM"
NORTH_AMERICA_KEYWORD_SHEET = "북미키워드"
DEFAULT_NORTH_AMERICA_KEYWORDS = [
    "AMQ Soluti",
    "BIF NY",
    "CLEVER",
    "12FAM",
    "FW",
    "GUAM",
    "12IMK3",
    "INICS",
    "BIF LA",
    "12LCQ",
    "12LQK",
    "MAVINS",
    "FURSYS NJ",
    "12ONK",
    "12SDA",
    "TOF",
    "JIMP",
]
WORKSHEET_PREFIXES = {
    "order_lines": "수주내역정보_",
    "management": "수주관리_",
    "progress": "공정 진행정보_",
    "inventory": "재고현황_",
}


ORDERS = [
    {
        "id": "ORD-001",
        "type": "내수",
        "office": "대리점기타",
        "customer": "(주)이그니스 10층",
        "displayName": "(주)이그니스 10층",
        "secondaryName": "퍼티션 추가",
        "site": "서울특별시 성동구 성수일로6길 53",
        "title": "퍼티션 추가",
        "startDate": "2026-03-20",
        "endDate": "2026-03-20",
        "items": 2,
        "groupedOrders": 39,
        "amount": 520,
        "category": "가구",
        "color": "061",
        "code": "XLFA10013P",
        "risk": "보통",
    },
    {
        "id": "ORD-002",
        "type": "내수",
        "office": "대리점기타",
        "customer": "(주)이그니스 10층",
        "displayName": "(주)이그니스 10층",
        "secondaryName": "업무석 20인 추가",
        "site": "서울특별시 성동구 성수일로6길 53",
        "title": "업무석 20인 추가",
        "startDate": "2026-03-20",
        "endDate": "2026-03-20",
        "items": 8,
        "groupedOrders": 8,
        "amount": 430,
        "category": "책상",
        "color": "WT",
        "code": "DESK2026A",
        "risk": "높음",
    },
    {
        "id": "ORD-003",
        "type": "수출",
        "office": "해외사업",
        "customer": "ABC Global Office",
        "displayName": "ABC Global Office",
        "secondaryName": "캐비닛 30EA 출하",
        "site": "부산항 물류센터",
        "title": "캐비닛 30EA 출하",
        "startDate": "2026-03-24",
        "endDate": "2026-03-24",
        "items": 1,
        "groupedOrders": 18,
        "amount": 780,
        "category": "캐비닛",
        "color": "BK",
        "code": "CAB9002",
        "risk": "보통",
    },
    {
        "id": "ORD-004",
        "type": "내수",
        "office": "수도권",
        "customer": "현대 오피스 리뉴얼",
        "displayName": "현대 오피스 리뉴얼",
        "secondaryName": "회의실 가구 납품",
        "site": "판교 제2테크노밸리",
        "title": "회의실 가구 납품",
        "startDate": "2026-03-25",
        "endDate": "2026-03-25",
        "items": 3,
        "groupedOrders": 12,
        "amount": 690,
        "category": "회의실",
        "color": "GY",
        "code": "MEET3310",
        "risk": "낮음",
    },
    {
        "id": "ORD-005",
        "type": "수출",
        "office": "해외사업",
        "customer": "Tokyo Branch",
        "displayName": "Tokyo Branch",
        "secondaryName": "데스크 50EA 선적",
        "site": "인천항",
        "title": "데스크 50EA 선적",
        "startDate": "2026-03-27",
        "endDate": "2026-03-28",
        "items": 5,
        "groupedOrders": 27,
        "amount": 1210,
        "category": "책상",
        "color": "NA",
        "code": "EXP5007",
        "risk": "높음",
    },
    {
        "id": "ORD-006",
        "type": "내수",
        "office": "충청권",
        "customer": "세종 스마트센터",
        "displayName": "세종 스마트센터",
        "secondaryName": "로비 가구 교체",
        "site": "세종시 도움8로",
        "title": "로비 가구 교체",
        "startDate": "2026-03-30",
        "endDate": "2026-03-30",
        "items": 4,
        "groupedOrders": 16,
        "amount": 410,
        "category": "로비",
        "color": "IV",
        "code": "LOB2211",
        "risk": "보통",
    },
]

ORDER_ITEMS = {
    "ORD-001": [
        {"name": "LF0102-협탁", "spec": "W1000", "color": "061", "qty": 12},
        {"name": "퍼티션 패널", "spec": "H1200", "color": "LG", "qty": 18},
        {"name": "케이블 트레이", "spec": "기본형", "color": "BK", "qty": 9},
    ],
    "ORD-002": [
        {"name": "업무용 책상", "spec": "1600x800", "color": "WT", "qty": 20},
        {"name": "사이드 서랍", "spec": "3단", "color": "GY", "qty": 20},
        {"name": "스크린 패널", "spec": "1200", "color": "NV", "qty": 20},
    ],
    "ORD-003": [
        {"name": "캐비닛", "spec": "5단", "color": "BK", "qty": 30},
        {"name": "잠금장치", "spec": "실린더형", "color": "SV", "qty": 30},
    ],
    "ORD-004": [
        {"name": "회의 테이블", "spec": "8인용", "color": "GY", "qty": 2},
        {"name": "회의 의자", "spec": "메쉬형", "color": "BK", "qty": 16},
        {"name": "보조장", "spec": "하부장", "color": "WT", "qty": 3},
    ],
    "ORD-005": [
        {"name": "데스크", "spec": "1400x700", "color": "NA", "qty": 50},
        {"name": "모니터 암", "spec": "싱글", "color": "BK", "qty": 50},
        {"name": "케이블덕트", "spec": "수평형", "color": "SV", "qty": 50},
    ],
    "ORD-006": [
        {"name": "로비 테이블", "spec": "원형", "color": "IV", "qty": 4},
        {"name": "소파", "spec": "3인용", "color": "BG", "qty": 2},
        {"name": "가이드 데스크", "spec": "안내형", "color": "WT", "qty": 1},
    ],
}

WEEKDAY_LABELS = ["일", "월", "화", "수", "목", "금", "토"]

METRIC_DETAILS = {
    "totalOrders": {
        "title": "주요 수주건 세부내역",
        "description": "현재 필터 기준으로 집계된 주요 수주건 목록입니다.",
    },
    "groupedCount": {
        "title": "통합 수주건 수 세부내역",
        "description": "통합 처리된 수주건 수량과 대표 수주건별 묶음 현황입니다.",
    },
    "itemCount": {
        "title": "주제별 정리 항목 수 세부내역",
        "description": "품목, 코드, 색상, 카테고리 기준으로 정리된 항목입니다.",
    },
    "totalAmount": {
        "title": "단위작업 누적 규모 세부내역",
        "description": "선택 월에 집계된 수주건의 누적 작업 규모입니다.",
    },
}

TYPE_COLORS = {
    "내수": {"bg": "#e0f2fe", "text": "#0369a1", "border": "#bae6fd"},
    "수출": {"bg": "#f3e8ff", "text": "#7e22ce", "border": "#e9d5ff"},
}

RISK_COLORS = {
    "높음": {"bg": "#fee2e2", "text": "#b91c1c", "border": "#fecaca"},
    "보통": {"bg": "#fef3c7", "text": "#b45309", "border": "#fde68a"},
    "낮음": {"bg": "#dcfce7", "text": "#15803d", "border": "#bbf7d0"},
}

NA_COLORS = {"bg": "#ffedd5", "text": "#9a3412", "border": "#fdba74"}

GOOGLE_SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]


def today_kst() -> date:
    return datetime.now(ZoneInfo("Asia/Seoul")).date()


def inject_css():
    st.markdown(
        """
        <style>
        .stApp {
            background: #f8fafc;
        }
        .block-container {
            padding-top: 1.6rem;
            padding-bottom: 2rem;
            max-width: 98vw;
            width: 98vw;
        }
        div[data-testid="stMetric"] {
            background: white;
            border-radius: 24px;
            padding: 18px 20px;
            box-shadow: 0 1px 3px rgba(15, 23, 42, 0.08);
            border: 1px solid #e2e8f0;
        }
        .hero-card, .soft-card, .section-card {
            background: white;
            border-radius: 26px;
            box-shadow: 0 1px 3px rgba(15, 23, 42, 0.08);
            border: 1px solid #e2e8f0;
        }
        .hero-card {
            padding: 28px;
            margin-bottom: 18px;
        }
        .soft-card {
            background: #f8fafc;
            padding: 16px 18px;
            min-height: 88px;
        }
        .section-card {
            padding: 20px;
        }
        div[data-testid="stVerticalBlockBorderWrapper"] {
            background: white;
            border-radius: 20px;
            border-color: #e2e8f0 !important;
            box-shadow: 0 1px 2px rgba(15, 23, 42, 0.06);
            padding: 12px 14px;
        }
        div[data-testid="stVerticalBlockBorderWrapper"] [data-testid="stMarkdownContainer"] h2 {
            margin-top: 0;
        }
        .badge {
            display: inline-block;
            padding: 4px 10px;
            border-radius: 999px;
            font-size: 12px;
            font-weight: 600;
            border: 1px solid transparent;
            margin-right: 6px;
            margin-bottom: 6px;
        }
        .calendar-grid {
            border: 1px solid #e2e8f0;
            border-radius: 22px;
            overflow: hidden;
            background: white;
        }
        .calendar-head {
            background: #f8fafc;
            font-size: 13px;
            font-weight: 700;
            color: #334155;
            padding: 12px 10px;
            border-bottom: 1px solid #e2e8f0;
            text-align: left;
        }
        .calendar-head.sun, .calendar-daynum.sun {
            color: #ef4444;
        }
        .calendar-head.sat, .calendar-daynum.sat {
            color: #2563eb;
        }
        .calendar-cell {
            min-height: 132px;
            padding: 8px;
            border-right: 1px solid #e2e8f0;
            border-bottom: 1px solid #e2e8f0;
            background: white;
        }
        .calendar-cell.selected {
            background: #fef3c7;
        }
        .calendar-daynum {
            font-size: 13px;
            margin-bottom: 8px;
        }
        .detail-box {
            background: #f8fafc;
            border-radius: 18px;
            padding: 14px 16px;
        }
        .order-card {
            border-radius: 24px;
            border: 1px solid #e2e8f0;
            background: white;
            padding: 18px;
            box-shadow: 0 1px 2px rgba(15, 23, 42, 0.06);
            min-height: 230px;
        }
        .order-card.selected {
            border-color: #0f172a;
            background: #f8fafc;
        }
        .tiny {
            color: #94a3b8;
            font-size: 12px;
        }
        .table-wrap {
            overflow: hidden;
            border: 1px solid #e2e8f0;
            border-radius: 18px;
            background: white;
        }
        .subtle-title {
            font-size: 15px;
            color: #64748b;
            margin-bottom: 4px;
        }
        .summary-card {
            background: white;
            border: 1px solid #e2e8f0;
            border-radius: 18px;
            padding: 16px 18px;
            box-shadow: 0 1px 2px rgba(15, 23, 42, 0.06);
        }
        .summary-title {
            font-size: 20px;
            color: #334155;
            font-weight: 700;
            margin-bottom: 10px;
            line-height: 1.25;
        }
        .summary-value {
            font-size: 32px;
            color: #0f172a;
            font-weight: 800;
            line-height: 1.1;
        }
        .summary-sub {
            margin-top: 8px;
            font-size: 12px;
            color: #64748b;
        }
        .summary-subtitle {
            font-size: 18px;
            color: #334155;
            font-weight: 700;
            margin-bottom: 10px;
            line-height: 1.25;
        }
        .summary-total {
            font-size: 38px;
            color: #0f172a;
            font-weight: 800;
            line-height: 1.1;
        }
        .summary-partial {
            margin-top: 8px;
            font-size: 13px;
            color: #64748b;
            font-weight: 600;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def month_label(month_str: str) -> str:
    year, month = month_str.split("-")
    return f"{year}년 {int(month)}월"


def format_korean_date(date_str: str) -> str:
    y, m, d = date_str.split("-")
    return f"{y}년 {int(m)}월 {int(d)}일"


def extract_snapshot_datetime_from_title(title: str) -> datetime | None:
    text = str(title or "").strip()
    if not text:
        return None

    patterns = [
        r"(20\d{2})[-/.](\d{2})[-/.](\d{2})[_\-\s](\d{2})(\d{2})(\d{2})?",
        r"(20\d{2})(\d{2})(\d{2})[_\-\s]?(\d{2})(\d{2})(\d{2})",
        r"(20\d{2})[-/.](\d{2})[-/.](\d{2})[\sT](\d{2}):(\d{2})(?::(\d{2}))?",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if not match:
            continue
        groups = match.groups()
        if len(groups) == 6 and groups[5] is None:
            groups = (*groups[:5], "00")
        groups = tuple("00" if g is None else g for g in groups)
        try:
            y, m, d, hh, mm, ss = [int(v) for v in groups[:6]]
            return datetime(y, m, d, hh, mm, ss)
        except ValueError:
            continue

    date_match = re.search(r"(20\d{2})(\d{2})(\d{2})", text)
    if date_match:
        try:
            y, m, d = [int(v) for v in date_match.groups()]
            return datetime(y, m, d, 0, 0, 0)
        except ValueError:
            return None
    return None


def build_source_snapshot_label(source_titles: dict) -> str:
    if not isinstance(source_titles, dict):
        return "기준 시각 확인 불가"
    parsed = []
    for title in source_titles.values():
        dt = extract_snapshot_datetime_from_title(str(title))
        if dt:
            parsed.append(dt)
    if not parsed:
        return "기준 시각 확인 불가"
    latest_dt = max(parsed)
    return latest_dt.strftime("%Y-%m-%d %H:%M")


def format_korean_amount_unit(amount: int) -> str:
    value = max(0, int(amount or 0))
    eok = value // 100_000_000
    chunman = (value % 100_000_000) // 10_000_000
    man = (value % 10_000_000) // 10_000

    parts = []
    if eok:
        parts.append(f"{eok}억")
    if chunman:
        parts.append(f"{chunman}천만")
    if not eok and not chunman:
        if man:
            parts.append(f"{man}만")
        else:
            parts.append("0")
    return " ".join(parts)


def clip_text(text: str, limit: int = 22) -> str:
    value = str(text)
    if len(value) <= limit:
        return value
    return value[: limit - 1] + "…"


def first_nonempty(series: pd.Series, fallback: str = "") -> str:
    for value in series:
        text = str(value).strip()
        if text and text.lower() != "nan":
            return text
    return fallback


def first_nonempty_values(values, fallback: str = "") -> str:
    for value in values:
        text = str(value).strip()
        if text and text.lower() != "nan":
            return text
    return fallback


def normalize_order_no(value) -> str:
    if value is None:
        return ""
    return "".join(ch for ch in str(value).strip() if ch.isalnum()).upper()


def normalize_address(value) -> str:
    text = str(value or "").strip()
    if not text or text.lower() == "nan":
        return ""
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"(지하\s*\d+층|지상\s*\d+층|\d+\s*층|\d+\s*F|\d+F)\b", "", text, flags=re.IGNORECASE)
    text = re.sub(r"(\d+\s*호|\d+\s*실|[A-Z]동\s*\d+호|[A-Z]동|\d+동)\b", "", text, flags=re.IGNORECASE)
    text = re.sub(r"(회의실|사무실|창고|현장|센터|로비|데스크)\s*$", "", text)
    text = re.sub(r"[,\-]+$", "", text).strip()
    text = re.sub(r"\s{2,}", " ", text)
    return text


def simplify_project_name(value) -> str:
    text = str(value or "").strip()
    if not text or text.lower() == "nan":
        return ""
    text = re.sub(r"\([^)]*(주문품|추가|변경|납품|설치|시공|별도)[^)]*\)", "", text)
    text = re.sub(r"\b(주문품|추가|변경|납품|설치|시공|케비넷|캐비넷|책상|의자|가구|스크린|퍼티션|데스크)\b.*$", "", text)
    text = re.sub(r"[-_/|]+.*$", "", text)
    text = re.sub(r"\s+", " ", text).strip(" -_/|")
    return text


def normalize_project_key(value) -> str:
    text = simplify_project_name(value)
    text = re.sub(r"[^0-9A-Za-z가-힣]", "", text)
    return text.upper()


def shorten_item_name_for_display(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    parts = text.split("-")
    if len(parts) >= 2:
        code = parts[0].strip()
        change_type = parts[1].strip()
        # Strip bracket suffixes and keep first 2 chars of change type.
        change_type = re.sub(r"\[.*$", "", change_type).strip()
        short_change = change_type[:2] if change_type else ""
        if code and short_change:
            return f"{code}-{short_change}"
    return text


def with_row_no_and_total(
    df: pd.DataFrame,
    qty_candidates: tuple[str, ...] = ("수주량", "수량", "qty"),
) -> pd.DataFrame:
    if df is None:
        return pd.DataFrame()
    out = df.copy()
    if out.empty:
        return out
    out = out.reset_index(drop=True)
    out.insert(0, "번호", range(1, len(out) + 1))

    qty_col = next((col for col in qty_candidates if col in out.columns), None)
    if not qty_col:
        return out

    qty_sum = pd.to_numeric(out[qty_col], errors="coerce").fillna(0).sum()
    if float(qty_sum).is_integer():
        qty_sum_value = int(qty_sum)
    else:
        qty_sum_value = float(qty_sum)

    total_row = {col: "" for col in out.columns}
    total_row["번호"] = "합계"
    total_row[qty_col] = qty_sum_value
    out = pd.concat([out, pd.DataFrame([total_row])], ignore_index=True)
    return out


def format_number_value(value):
    if value is None:
        return ""
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return ""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        num = float(value)
        if num.is_integer():
            return f"{int(num):,}"
        return f"{num:,.2f}".rstrip("0").rstrip(".")
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return text
        normalized = text.replace(",", "")
        if re.fullmatch(r"-?\d+(\.\d+)?", normalized):
            try:
                num = float(normalized)
                if num.is_integer():
                    return f"{int(num):,}"
                return f"{num:,.2f}".rstrip("0").rstrip(".")
            except Exception:
                return value
    return value


def format_df_numbers_for_display(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    out = df.copy()
    for col in out.columns:
        out[col] = out[col].map(format_number_value)
    return out


def dataframe_to_styled_excel_bytes(df: pd.DataFrame, sheet_name: str = "통합품목") -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.book[sheet_name]

        body_font = Font(name="Malgun Gothic", size=11, bold=False)
        head_font = Font(name="Malgun Gothic", size=12, bold=True)
        center = Alignment(horizontal="center", vertical="center")
        header_fill = PatternFill(fill_type="solid", fgColor="EDEDED")
        thin_side = Side(style="thin", color="000000")
        border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        max_row = ws.max_row
        max_col = ws.max_column
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = center
                cell.border = border
                if row == 1:
                    cell.font = head_font
                    cell.fill = header_fill
                else:
                    cell.font = body_font
                    cell.fill = PatternFill(fill_type=None)

        for col_idx in range(1, max_col + 1):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            max_length = 0
            for row_idx in range(1, max_row + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is None:
                    continue
                max_length = max(max_length, len(str(value)))
            ws.column_dimensions[col_letter].width = min(max(12, max_length + 2), 48)

    return output.getvalue()


def build_display_name(title, fallback: str = "") -> str:
    text = str(title or "").strip()
    fallback_text = str(fallback or "").strip()
    if not text:
        return fallback_text

    text = re.sub(r"\([^)]*주문품[^)]*\)", "", text)
    text = text.split("_")[0].strip()
    parts = [part.strip() for part in re.split(r"\s*-\s*", text) if part.strip()]
    stop_pattern = re.compile(r"(층|호|실|주문품|추가|변경|납품|설치|사이드|하부장|상부장|책상|의자|가구|스크린|퍼티션|데스크)")

    if len(parts) >= 2:
        kept = [parts[0]]
        for part in parts[1:]:
            if stop_pattern.search(part):
                break
            part = re.sub(r"(?i)\b(BIFC)\d+\b", r"\1", part)
            kept.append(part)
            if len(kept) >= 2:
                break
        text = "-".join(kept)

    text = re.sub(r"(?i)\b(BIFC)\d+\b", r"\1", text)
    text = re.sub(r"\s+", " ", text).strip(" -_/|")
    return text or fallback_text


def compile_keyword_regex(keywords: list[str]) -> str:
    normalized_keywords = []
    for kw in keywords:
        normalized_kw = normalize_match_text(kw)
        if normalized_kw:
            normalized_keywords.append(re.escape(normalized_kw))
    if not normalized_keywords:
        return ""
    return "|".join(sorted(set(normalized_keywords), key=len, reverse=True))


def record_perf_step(perf_rows: list[dict], step: str, start_time: float) -> float:
    now = perf_counter()
    perf_rows.append({"step": step, "ms": round((now - start_time) * 1000, 1)})
    return now


@lru_cache(maxsize=50000)
def extract_name_tokens(value: str) -> frozenset[str]:
    text = build_display_name(value, "")
    if not text:
        return frozenset()
    raw_tokens = re.split(r"[\s\-_()/]+", text)
    stopwords = {
        "주",
        "주식회사",
        "추가",
        "변경",
        "납품",
        "설치",
        "시공",
        "주문품",
        "사무비품",
        "비규격",
    }
    tokens: set[str] = set()
    for token in raw_tokens:
        token = token.strip()
        if len(token) < 2:
            continue
        if token in stopwords:
            continue
        tokens.add(token.upper())
    return frozenset(tokens)


def name_similarity(a, b) -> float:
    a_tokens = extract_name_tokens(a)
    b_tokens = extract_name_tokens(b)
    if not a_tokens or not b_tokens:
        return 0.0
    inter = len(a_tokens & b_tokens)
    union = len(a_tokens | b_tokens)
    return inter / union if union else 0.0


def to_numeric(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series.astype(str).str.replace(",", "", regex=False), errors="coerce").fillna(0)


@st.cache_resource(show_spinner=False)
def open_spreadsheet():
    creds = None

    def _to_plain_dict(value):
        if isinstance(value, Mapping):
            return {k: _to_plain_dict(v) for k, v in value.items()}
        return value

    def _normalize_service_account_info(info: dict) -> dict:
        normalized = dict(info)
        private_key = str(normalized.get("private_key", ""))
        # secrets에 \n 문자열로 저장된 경우 실제 줄바꿈으로 복원
        if "\\n" in private_key:
            normalized["private_key"] = private_key.replace("\\n", "\n")
        return normalized

    # 1) Streamlit Cloud secrets 우선
    parse_errors: list[str] = []
    try:
        try:
            top_secrets = st.secrets.to_dict()
        except Exception:
            top_secrets = {k: st.secrets[k] for k in st.secrets.keys()}

        # A. 중첩 객체 방식: [gcp_service_account] ... (권장)
        for key in ["gcp_service_account", "GCP_SERVICE_ACCOUNT", "google_service_account", "service_account"]:
            raw = top_secrets.get(key)
            if not raw:
                continue
            try:
                if isinstance(raw, str):
                    info = json.loads(raw)
                else:
                    info = _to_plain_dict(raw)
                info = _normalize_service_account_info(info)
                creds = Credentials.from_service_account_info(info, scopes=GOOGLE_SCOPES)
                break
            except Exception as exc:
                parse_errors.append(f"{key}: {exc}")

        # B. 평면 키 방식(type/project_id/private_key...)
        if creds is None:
            required = ["type", "project_id", "private_key", "client_email", "token_uri"]
            if all(k in top_secrets for k in required):
                try:
                    info = {k: top_secrets.get(k) for k in required + [
                        "private_key_id", "client_id", "auth_uri",
                        "auth_provider_x509_cert_url", "client_x509_cert_url", "universe_domain"
                    ]}
                    info = _normalize_service_account_info(info)
                    creds = Credentials.from_service_account_info(info, scopes=GOOGLE_SCOPES)
                except Exception as exc:
                    parse_errors.append(f"flat_keys: {exc}")

        # C. JSON 문자열 방식
        if creds is None:
            for json_key in ["GCP_SERVICE_ACCOUNT_JSON", "gcp_service_account_json", "GOOGLE_SERVICE_ACCOUNT_JSON"]:
                if json_key not in top_secrets:
                    continue
                try:
                    info = json.loads(str(top_secrets[json_key]))
                    info = _normalize_service_account_info(info)
                    creds = Credentials.from_service_account_info(info, scopes=GOOGLE_SCOPES)
                    break
                except Exception as exc:
                    parse_errors.append(f"{json_key}: {exc}")
    except Exception as exc:
        parse_errors.append(f"secrets_read: {exc}")

    # 1-2) 환경변수 fallback (Streamlit Cloud Advanced settings에서 설정 가능)
    if creds is None:
        for env_key in ["GCP_SERVICE_ACCOUNT_JSON", "GOOGLE_SERVICE_ACCOUNT_JSON"]:
            raw_env = os.getenv(env_key, "").strip()
            if not raw_env:
                continue
            try:
                info = json.loads(raw_env)
                info = _normalize_service_account_info(info)
                creds = Credentials.from_service_account_info(info, scopes=GOOGLE_SCOPES)
                break
            except Exception as exc:
                parse_errors.append(f"env:{env_key}: {exc}")

    # 2) 로컬 파일 fallback
    if creds is None:
        if not GOOGLE_CREDENTIALS_FILE.exists():
            parse_msg = (" / ".join(parse_errors)) if parse_errors else "no-credentials"
            raise FileNotFoundError(
                "구글 인증정보를 찾지 못했습니다. "
                "Streamlit Cloud의 Secrets에 [gcp_service_account] 블록을 넣거나 "
                "GCP_SERVICE_ACCOUNT_JSON 문자열을 설정해 주세요. "
                f"(parse detail: {parse_msg})"
            )
        creds = Credentials.from_service_account_file(str(GOOGLE_CREDENTIALS_FILE), scopes=GOOGLE_SCOPES)

    client = gspread.authorize(creds)
    return client.open_by_key(GOOGLE_SPREADSHEET_ID)


def normalize_match_text(value: str) -> str:
    text = str(value or "").upper()
    return re.sub(r"[^A-Z0-9가-힣]", "", text)


def find_matching_keywords(text: str, keywords: list[str]) -> list[str]:
    normalized_text = normalize_match_text(text)
    if not normalized_text:
        return []
    matches: list[str] = []
    for raw_kw in keywords:
        kw = str(raw_kw or "").strip()
        if not kw:
            continue
        normalized_kw = normalize_match_text(kw)
        if normalized_kw and normalized_kw in normalized_text:
            matches.append(kw)
    return matches


def ensure_north_america_keyword_worksheet(spreadsheet):
    try:
        return spreadsheet.worksheet(NORTH_AMERICA_KEYWORD_SHEET)
    except WorksheetNotFound:
        worksheet = spreadsheet.add_worksheet(title=NORTH_AMERICA_KEYWORD_SHEET, rows=400, cols=2)
        initial_rows = [["키워드", "활성"]] + [[kw, "Y"] for kw in DEFAULT_NORTH_AMERICA_KEYWORDS]
        worksheet.update(range_name="A1", values=initial_rows)
        return worksheet


def load_north_america_keywords(spreadsheet) -> tuple[list[dict], list[str]]:
    worksheet = ensure_north_america_keyword_worksheet(spreadsheet)
    values = worksheet.get_all_values()
    if not values:
        rows = [{"키워드": kw, "활성": True} for kw in DEFAULT_NORTH_AMERICA_KEYWORDS]
        return rows, DEFAULT_NORTH_AMERICA_KEYWORDS.copy()

    header = [str(v).strip() for v in values[0]]
    data_rows = values[1:]
    idx_kw = header.index("키워드") if "키워드" in header else 0
    idx_on = header.index("활성") if "활성" in header else None

    rows: list[dict] = []
    for row in data_rows:
        keyword = row[idx_kw].strip() if len(row) > idx_kw else ""
        if not keyword:
            continue
        raw_on = row[idx_on].strip() if idx_on is not None and len(row) > idx_on else "Y"
        active = str(raw_on).upper() in {"Y", "YES", "TRUE", "1", "활성", "사용"}
        rows.append({"키워드": keyword, "활성": active})

    if not rows:
        rows = [{"키워드": kw, "활성": True} for kw in DEFAULT_NORTH_AMERICA_KEYWORDS]

    active_keywords = [r["키워드"] for r in rows if r.get("활성")]
    return rows, active_keywords


def save_north_america_keywords(rows: list[dict]):
    spreadsheet = open_spreadsheet()
    worksheet = ensure_north_america_keyword_worksheet(spreadsheet)
    sanitized: list[list[str]] = [["키워드", "활성"]]
    for row in rows:
        keyword = str(row.get("키워드", "")).strip()
        if not keyword:
            continue
        active = bool(row.get("활성", True))
        sanitized.append([keyword, "Y" if active else "N"])
    if len(sanitized) == 1:
        sanitized += [[kw, "Y"] for kw in DEFAULT_NORTH_AMERICA_KEYWORDS]
    worksheet.clear()
    worksheet.update(range_name="A1", values=sanitized)


def worksheet_to_df(worksheet) -> pd.DataFrame:
    values = worksheet.get_all_values()
    if not values:
        return pd.DataFrame()
    header = values[0]
    rows = values[1:]
    padded = [row + [""] * (len(header) - len(row)) for row in rows]
    df = pd.DataFrame(padded, columns=header)
    return df.fillna("")


def normalize_dataframe_columns(df: pd.DataFrame) -> pd.DataFrame:
    renamed = {}
    for col in df.columns:
        normalized = str(col).strip()
        normalized = re.sub(r"[▲▼]", "", normalized)
        normalized = re.sub(r"\s+", "", normalized)
        renamed[col] = normalized
    return df.rename(columns=renamed)


def select_columns_with_defaults(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    selected = df.reindex(columns=columns).copy()
    return selected.fillna("")


def _find_next_days(start_day: date, count: int = 7) -> list[date]:
    return [start_day + timedelta(days=offset) for offset in range(1, count + 1)]


def compute_inventory_d5_metrics(df: pd.DataFrame) -> tuple[pd.Series, pd.Series]:
    if df.empty:
        return pd.Series(dtype=float), pd.Series(dtype=float)

    date_col_info: list[tuple[int, str, date]] = []
    for idx, col in enumerate(df.columns):
        try:
            day = date.fromisoformat(str(col))
            date_col_info.append((idx, col, day))
        except Exception:
            continue
    if not date_col_info:
        empty = pd.Series([math.nan] * len(df), index=df.index, dtype=float)
        return empty, empty

    first_row = df.iloc[0] if len(df) > 0 else pd.Series(dtype=object)
    markers = {"출예", "물입예", "예량"}
    first_row_marker_count = 0
    if not first_row.empty:
        first_row_marker_count = int(
            sum(str(v).strip() in markers for v in first_row.values)
        )
    marker_row_exists = first_row_marker_count >= 6

    # D+7 구간은 비근무일 포함한 달력일 기준으로 본다.
    target_days = _find_next_days(today_kst(), count=7)
    target_day_set = set(target_days)
    selected_date_cols = [info for info in date_col_info if info[2] in target_day_set]
    if not selected_date_cols:
        empty = pd.Series([math.nan] * len(df), index=df.index, dtype=float)
        return empty, empty

    qty_cols: list[str] = []
    out_cols: list[str] = []
    columns_list = list(df.columns)
    for col_idx, date_col, _ in selected_date_cols:
        qty_col = date_col
        out_col = date_col
        if marker_row_exists:
            candidates = []
            for shift in [0, 1, 2, 3]:
                pos = col_idx + shift
                if pos >= len(columns_list):
                    continue
                cand = columns_list[pos]
                marker_text = str(first_row.get(cand, "")).strip()
                candidates.append((cand, marker_text))
            matched = next((cand for cand, marker in candidates if marker == "예량"), None)
            if matched is not None:
                qty_col = matched
            matched_out = next((cand for cand, marker in candidates if marker == "출예"), None)
            if matched_out is not None:
                out_col = matched_out
        qty_cols.append(qty_col)
        out_cols.append(out_col)

    work_df = df.iloc[1:].copy() if marker_row_exists and len(df) > 1 else df.copy()
    if work_df.empty:
        empty = pd.Series([math.nan] * len(df), index=df.index, dtype=float)
        return empty, empty

    for col in qty_cols:
        if col not in work_df.columns:
            work_df[col] = ""
    for col in out_cols:
        if col not in work_df.columns:
            work_df[col] = ""
    qty_frame = work_df[qty_cols].apply(to_numeric)
    min_series = qty_frame.min(axis=1, skipna=True)
    out_frame = work_df[out_cols].apply(to_numeric)
    out_sum_series = out_frame.sum(axis=1, skipna=True)

    aligned = pd.Series([math.nan] * len(df), index=df.index, dtype=float)
    aligned.loc[work_df.index] = min_series
    aligned_out = pd.Series([math.nan] * len(df), index=df.index, dtype=float)
    aligned_out.loc[work_df.index] = out_sum_series
    return aligned, aligned_out


@st.cache_resource(ttl=300, show_spinner="스프레드시트 데이터를 불러오는 중입니다...")
def load_dashboard_base_data():
    perf_rows: list[dict] = []
    step_start = perf_counter()
    spreadsheet = open_spreadsheet()
    worksheets = spreadsheet.worksheets()
    na_keyword_rows, na_active_keywords = load_north_america_keywords(spreadsheet)
    step_start = record_perf_step(perf_rows, "sheet_open_and_keyword_load", step_start)

    latest_titles: dict[str, str] = {}
    frames: dict[str, pd.DataFrame] = {}
    for key, prefix in WORKSHEET_PREFIXES.items():
        matches = [ws for ws in worksheets if ws.title.startswith(prefix)]
        if not matches:
            if key == "inventory":
                latest_titles[key] = "(미연결)"
                frames[key] = pd.DataFrame(columns=["단품코드", "색상", "현재고", "재고금액", "기간총입고", "기간총출고"])
                continue
            if key == "progress":
                latest_titles[key] = "(미연결)"
                frames[key] = pd.DataFrame(columns=["단품코드", "색상", "계획", "생산", "잔량", "진행률", "진행상태", "관리번호"])
                continue
            raise ValueError(f"'{prefix}'로 시작하는 시트를 찾지 못했습니다.")
        latest = sorted(matches, key=lambda ws: ws.title)[-1]
        latest_titles[key] = latest.title
        frames[key] = normalize_dataframe_columns(worksheet_to_df(latest))
    step_start = record_perf_step(perf_rows, "worksheet_select_and_to_dataframe", step_start)

    order_lines = select_columns_with_defaults(
        frames["order_lines"],
        [
            "수주번호",
            "수주건명",
            "브랜드",
            "브랜드명",
            "사업소",
            "확정납기",
            "단품코드",
            "색상",
            "수주량",
            "수주금액",
            "제품구분",
            "재고구분",
            "표준구분",
            "단품명칭",
        ],
    )
    management = select_columns_with_defaults(
        frames["management"],
        [
            "수주번호",
            "수주건명",
            "브랜드",
            "브랜드명",
            "영업건명",
            "납품처주소",
            "사업소",
            "시공센터",
            "시공유무",
            "확정납기",
            "대리점",
            "실적대리점",
        ],
    )
    progress = select_columns_with_defaults(
        frames["progress"],
        ["단품코드", "색상", "계획", "생산", "잔량", "진행률", "진행상태", "관리번호"],
    )
    inventory = frames["inventory"].copy()

    order_lines["수주번호_norm"] = order_lines["수주번호"].map(normalize_order_no)
    management["수주번호_norm"] = management["수주번호"].map(normalize_order_no)

    order_lines["확정납기_dt"] = pd.to_datetime(order_lines["확정납기"], errors="coerce")
    order_lines["수주량_num"] = to_numeric(order_lines["수주량"])
    order_lines["수주금액_num"] = to_numeric(order_lines.get("수주금액", pd.Series(dtype=object)))
    order_lines["line_key"] = (
        order_lines["단품코드"].astype(str).fillna("")
        + "||"
        + order_lines["색상"].astype(str).fillna("")
    )

    management_subset = management.drop_duplicates(subset=["수주번호_norm"], keep="first").copy()
    management_subset = management_subset.rename(
        columns={
            "수주번호": "관리수주번호",
            "수주건명": "관리수주건명",
            "브랜드": "관리브랜드",
            "브랜드명": "관리브랜드명",
            "영업건명": "관리영업건명",
            "사업소": "관리사업소",
            "시공센터": "관리시공센터",
            "시공유무": "관리시공유무",
            "확정납기": "관리확정납기",
        }
    )

    if not progress.empty:
        progress["계획_num"] = to_numeric(progress.get("계획", pd.Series(dtype=object)))
        progress["생산_num"] = to_numeric(progress.get("생산", pd.Series(dtype=object)))
        progress["잔량_num"] = to_numeric(progress.get("잔량", pd.Series(dtype=object)))
        progress["진행률_num"] = to_numeric(progress.get("진행률", pd.Series(dtype=object)))
        progress_agg = (
            progress.groupby(["단품코드", "색상"], dropna=False)
            .agg(
                계획=("계획_num", "sum"),
                생산=("생산_num", "sum"),
                잔량=("잔량_num", "sum"),
                평균진행률=("진행률_num", "max"),
                진행상태=("진행상태", lambda s: first_nonempty(s, "미확인")),
                관리번호수=("관리번호", "nunique"),
            )
            .reset_index()
        )
    else:
        progress_agg = pd.DataFrame(columns=["단품코드", "색상", "계획", "생산", "잔량", "평균진행률", "진행상태", "관리번호수"])

    if not inventory.empty:
        inventory = select_columns_with_defaults(
            inventory,
            list(inventory.columns),
        )
        d5_min_series, d5_out_sum_series = compute_inventory_d5_metrics(inventory)
        inventory["예량(D+7)_num"] = d5_min_series
        inventory["출고예정(D+7)_num"] = d5_out_sum_series
        inventory["현재고_num"] = to_numeric(inventory.get("현재고", pd.Series(dtype=object)))
        inventory["재고금액_num"] = to_numeric(inventory.get("재고금액", pd.Series(dtype=object)))
        inventory["기간총입고_num"] = to_numeric(inventory.get("기간총입고", pd.Series(dtype=object)))
        inventory["기간총출고_num"] = to_numeric(inventory.get("기간총출고", pd.Series(dtype=object)))
        inventory_agg = (
            inventory.groupby(["단품코드", "색상"], dropna=False)
            .agg(
                현재고=("현재고_num", "sum"),
                재고금액=("재고금액_num", "sum"),
                기간총입고=("기간총입고_num", "sum"),
                기간총출고=("기간총출고_num", "sum"),
                예량D7=("예량(D+7)_num", "min"),
                출고예정D7=("출고예정(D+7)_num", "sum"),
            )
            .reset_index()
        )
    else:
        inventory_agg = pd.DataFrame(
            columns=["단품코드", "색상", "현재고", "재고금액", "기간총입고", "기간총출고", "예량D7", "출고예정D7"]
        )

    merged = order_lines.merge(management_subset, on="수주번호_norm", how="left")
    merged = merged.merge(progress_agg, on=["단품코드", "색상"], how="left")
    merged = merged.merge(inventory_agg, on=["단품코드", "색상"], how="left")
    merged[["계획", "생산", "잔량", "평균진행률", "관리번호수"]] = merged[
        ["계획", "생산", "잔량", "평균진행률", "관리번호수"]
    ].fillna(0)
    merged[["현재고", "재고금액", "기간총입고", "기간총출고", "예량D7", "출고예정D7"]] = merged[
        ["현재고", "재고금액", "기간총입고", "기간총출고", "예량D7", "출고예정D7"]
    ].fillna(0)

    merged["대표수주건명"] = merged["관리영업건명"].replace("", pd.NA).fillna(
        merged["관리수주건명"].replace("", pd.NA)
    ).fillna(merged["수주건명"])
    merged["상세건명"] = merged["수주건명"].replace("", pd.NA).fillna(merged["대표수주건명"])
    merged["대표주소"] = merged["납품처주소"].replace("", pd.NA).fillna("주소 미등록")
    merged["기본주소"] = merged["대표주소"].map(normalize_address).replace("", "주소 미등록")
    merged["대표사업소"] = merged["관리사업소"].replace("", pd.NA).fillna(merged["사업소"])
    merged["구분"] = merged["대표사업소"].apply(lambda value: "수출" if "수출" in str(value).strip() else "내수")
    merged["대표프로젝트명"] = merged["대표수주건명"].map(simplify_project_name)
    merged["표시프로젝트명"] = merged["상세건명"].map(lambda value: build_display_name(value, ""))
    merged["프로젝트키"] = merged["대표수주건명"].map(normalize_project_key)
    merged["통합수주건키"] = ""
    brand_candidates = [col for col in ["브랜드", "브랜드명", "관리브랜드", "관리브랜드명"] if col in merged.columns]
    if brand_candidates:
        merged["브랜드표시"] = merged[brand_candidates].replace("", pd.NA).bfill(axis=1).iloc[:, 0].fillna("")
    else:
        merged["브랜드표시"] = ""
    merged["관리확정납기_dt"] = pd.to_datetime(merged.get("관리확정납기", pd.Series(dtype=object)), errors="coerce")
    merged["기준확정납기_dt"] = merged["확정납기_dt"].where(merged["확정납기_dt"].notna(), merged["관리확정납기_dt"])

    management_norm_col = next((col for col in management.columns if str(col).endswith("_norm")), None)
    merged_norm_col = next((col for col in merged.columns if str(col).endswith("_norm")), None)
    dealer_col = get_existing_column(management, ["대리점"])
    perf_dealer_col = get_existing_column(management, ["실적대리점"])
    if management_norm_col and merged_norm_col and (dealer_col or perf_dealer_col):
        map_cols = [management_norm_col]
        if dealer_col:
            map_cols.append(dealer_col)
        if perf_dealer_col:
            map_cols.append(perf_dealer_col)
        dealer_map = management[map_cols].copy().fillna("")
        agg_spec = {}
        if dealer_col:
            agg_spec[dealer_col] = "first"
        if perf_dealer_col:
            agg_spec[perf_dealer_col] = "first"
        dealer_map = dealer_map.groupby(management_norm_col, dropna=False).agg(agg_spec).reset_index()
        rename_cols = {management_norm_col: merged_norm_col}
        if dealer_col:
            rename_cols[dealer_col] = "_na_dealer"
        if perf_dealer_col:
            rename_cols[perf_dealer_col] = "_na_perf_dealer"
        dealer_map = dealer_map.rename(columns=rename_cols)
        merged = merged.merge(dealer_map, on=merged_norm_col, how="left")
    if "_na_dealer" not in merged.columns:
        merged["_na_dealer"] = ""
    if "_na_perf_dealer" not in merged.columns:
        merged["_na_perf_dealer"] = ""
    step_start = record_perf_step(perf_rows, "base_merge_and_enrichment", step_start)

    # Address first: create candidate groups by normalized site, then split them
    # by project-name similarity so unrelated jobs at the same site stay separate.
    for base_address, addr_group in merged.groupby("기본주소", dropna=False):
        if not base_address or base_address == "주소 미등록":
            merged.loc[addr_group.index, "통합수주건키"] = addr_group["수주번호_norm"]
            continue

        clusters: list[dict] = []
        for _, row in addr_group.sort_values(["기준확정납기_dt", "수주번호"]).iterrows():
            candidate_names = [
                row["표시프로젝트명"],
                row["대표프로젝트명"],
                row["상세건명"],
                row["대표수주건명"],
            ]
            candidate_name = first_nonempty_values(candidate_names, row["수주번호_norm"])
            candidate_display = build_display_name(candidate_name, row["수주번호_norm"])
            candidate_tokens = set(extract_name_tokens(candidate_name)) | set(extract_name_tokens(candidate_display))

            matched_cluster = None
            best_score = 0.0
            for cluster in clusters:
                cluster_tokens = cluster.get("tokens", set())
                if candidate_tokens and cluster_tokens and candidate_tokens.isdisjoint(cluster_tokens):
                    continue
                score = max(
                    name_similarity(candidate_name, cluster["name"]),
                    name_similarity(candidate_display, cluster["display"]),
                )
                if score > best_score:
                    best_score = score
                    matched_cluster = cluster

            if matched_cluster and best_score >= 0.35:
                matched_cluster["indices"].append(row.name)
                matched_cluster["tokens"] = set(matched_cluster.get("tokens", set())) | candidate_tokens
                if len(candidate_display) < len(matched_cluster["display"]) and candidate_display:
                    matched_cluster["display"] = candidate_display
                continue

            clusters.append(
                {
                    "indices": [row.name],
                    "name": candidate_name,
                    "display": candidate_display or row["수주번호_norm"],
                    "tokens": candidate_tokens,
                }
            )

        for cluster_idx, cluster in enumerate(clusters, start=1):
            cluster_key = f"{base_address}||{cluster_idx:02d}"
            merged.loc[cluster["indices"], "통합수주건키"] = cluster_key
    step_start = record_perf_step(perf_rows, "address_cluster_grouping", step_start)

    return {
        "merged": merged,
        "source_titles": latest_titles,
        "north_america_keyword_rows": na_keyword_rows,
        "north_america_active_keywords": na_active_keywords,
        "perf_base_ms": perf_rows,
    }


@st.cache_data(ttl=300, show_spinner="스프레드시트 데이터를 불러오는 중입니다...")
def load_dashboard_data(
    etc_amount_threshold: int = 100_000_000,
    product_family: str = "\ucda9\uc8fc",
    export_only_nonstock_custom: bool = True,
):
    perf_rows: list[dict] = []
    step_start = perf_counter()
    base_data = load_dashboard_base_data()
    step_start = record_perf_step(perf_rows, "base_data_fetch", step_start)
    latest_titles = base_data["source_titles"]
    na_keyword_rows = base_data["north_america_keyword_rows"]
    na_active_keywords = base_data["north_america_active_keywords"]
    merged = base_data["merged"].copy()
    perf_rows.extend(base_data.get("perf_base_ms", []))

    is_target_product = merged["제품구분"].isin(["충주1제품", "충주2제품", "충주상품", "충주2상품", "F우레탄제품"])
    if product_family == "\uc548\uc131":
        is_target_product = merged["제품구분"].astype(str).str.contains("\uc548\uc131", na=False)
    is_non_stock = merged["재고구분"].astype(str).str.strip() == "비재고"
    is_custom = merged["표준구분"].astype(str).str.contains("주문품", na=False)
    merged["주요후보"] = is_target_product & is_non_stock & is_custom

    candidate = merged[merged["주요후보"]].copy()
    candidate_item_agg = (
        candidate.groupby(["통합수주건키", "단품코드", "색상"], dropna=False)
        .agg(
            수량=("수주량_num", "sum"),
            품목명=("단품명칭", lambda s: first_nonempty(s, "품목명 없음")),
            제품구분=("제품구분", lambda s: first_nonempty(s, "미분류")),
            표준구분=("표준구분", lambda s: "주문품" if s.astype(str).str.contains("주문품", na=False).any() else "기타"),
            생산=("생산", "sum"),
            계획=("계획", "sum"),
            잔량=("잔량", "sum"),
            현재고=("현재고", "max"),
            재고금액=("재고금액", "max"),
            진행상태=("진행상태", lambda s: first_nonempty(s, "미확인")),
        )
        .reset_index()
    )
    candidate_item_agg["주요품목여부"] = (
        candidate_item_agg["표준구분"].astype(str).str.contains("주문품", na=False)
        & (candidate_item_agg["수량"] >= 30)
    )
    major_item_agg = candidate_item_agg[candidate_item_agg["주요품목여부"]].copy()

    major_group_keys = set(major_item_agg["통합수주건키"].unique())
    # OR 조건: 주문품 수주금액 합계가 3,000만원 이상이면 주요 수주건으로 인정
    order_amount_major_keys = set()
    order_amount_candidate = merged[
        merged["표준구분"].astype(str).str.contains("주문품", na=False)
        & is_target_product
    ].copy()
    if not order_amount_candidate.empty:
        order_amount_agg = (
            order_amount_candidate.groupby("통합수주건키", dropna=False)
            .agg(주문품수주금액합계=("수주금액_num", "sum"))
            .reset_index()
        )
        order_amount_major_keys = set(
            order_amount_agg.loc[
                order_amount_agg["주문품수주금액합계"] >= 30_000_000,
                "통합수주건키",
            ].astype(str)
        )
    major_group_keys.update(order_amount_major_keys)
    # 기타 후보: 표준품/주문품 무관 + 충주1/충주2/F우레탄 제품의 통합 수주금액 합계 1억 이상
    etc_group_keys = set()
    etc_amount_candidate = merged[is_target_product].copy()
    if not etc_amount_candidate.empty:
        etc_amount_agg = (
            etc_amount_candidate.groupby("통합수주건키", dropna=False)
            .agg(충주계열수주금액합계=("수주금액_num", "sum"))
            .reset_index()
        )
        etc_group_keys = set(
            etc_amount_agg.loc[
                etc_amount_agg["충주계열수주금액합계"] >= int(etc_amount_threshold),
                "통합수주건키",
            ].astype(str)
        )
    office_col_for_export = get_existing_column(merged, ["대표사업소", "사업소", "관리사업소"])
    product_col_for_export = get_existing_column(merged, ["제품구분"])
    standard_col_for_export = get_existing_column(merged, ["표준구분", "수지구분"])
    north_america_product_types = (
        {
            "충주1제품",
            "충주2제품",
            "충주상품",
            "충주2상품",
            "F우레탄제품",
            "베트남상품",
            "목제상품",
            "목제5상품",
            "목제6상품",
        }
        if product_family != "\uc548\uc131"
        else set()
    )
    if office_col_for_export:
        export_mask = merged[office_col_for_export].astype(str).str.contains("수출", na=False)
        if product_col_for_export:
            export_product_mask = is_target_product
        else:
            export_product_mask = pd.Series(False, index=merged.index)
        if export_only_nonstock_custom:
            export_target_mask = export_mask & export_product_mask & is_non_stock & is_custom
        else:
            export_target_mask = export_mask & export_product_mask
        export_group_keys = set(merged.loc[export_target_mask, "통합수주건키"].astype(str).unique())
        major_group_keys.update(export_group_keys)

        # 북미는 수출 후보와 완전히 분리해서 판정한다.
        # 조건: 수출사업소 + (대리점/실적대리점 북미키워드 매칭) + 북미 대상 제품구분
        if product_col_for_export:
            if product_family == "\uc548\uc131":
                north_america_product_mask = merged[product_col_for_export].astype(str).str.contains("\uc548\uc131", na=False)
            else:
                north_america_product_mask = merged[product_col_for_export].isin(north_america_product_types)
        else:
            north_america_product_mask = pd.Series(False, index=merged.index)
        dealer_text = (
            merged.get("_na_dealer", pd.Series("", index=merged.index)).astype(str)
            + " "
            + merged.get("_na_perf_dealer", pd.Series("", index=merged.index)).astype(str)
        )
        keyword_pattern = compile_keyword_regex(na_active_keywords)
        if keyword_pattern:
            normalized_dealer_text = dealer_text.map(normalize_match_text)
            north_america_keyword_mask = normalized_dealer_text.str.contains(keyword_pattern, regex=True, na=False)
        else:
            north_america_keyword_mask = pd.Series(False, index=merged.index)
        north_america_target_mask = export_mask & north_america_product_mask & north_america_keyword_mask
        north_america_group_keys = set(merged.loc[north_america_target_mask, "통합수주건키"].astype(str).unique())
        major_group_keys.update(north_america_group_keys)
    grouped_orders = merged[merged["통합수주건키"].isin(major_group_keys | etc_group_keys)].copy()
    step_start = record_perf_step(perf_rows, "major_etc_export_filtering", step_start)

    order_records: list[dict] = []
    items_by_order: dict[str, list[dict]] = {}
    related_by_order: dict[str, list[dict]] = {}
    detail_items_by_order: dict[str, list[dict]] = {}

    for group_key, group in grouped_orders.groupby("통합수주건키", dropna=False):
        major_items = major_item_agg[major_item_agg["통합수주건키"] == group_key].copy()
        if major_items.empty:
            item_name_col = get_existing_column(group, ["단품명칭", "품목명약칭", "품목명", "품목약칭"])
            product_class_col = get_existing_column(group, ["제품구분"])
            status_col = get_existing_column(group, ["진행상태"])
            fallback_group = group.copy()
            fallback_group["_item_name_fallback"] = (
                fallback_group[item_name_col].astype(str) if item_name_col else "품목명 없음"
            )
            fallback_group["_product_class_fallback"] = (
                fallback_group[product_class_col].astype(str) if product_class_col else "미분류"
            )
            fallback_group["_status_fallback"] = (
                fallback_group[status_col].astype(str) if status_col else "미확인"
            )
            major_items = (
                fallback_group.groupby(["단품코드", "색상"], dropna=False)
                .agg(
                    수량=("수주량_num", "sum"),
                    품목명=("_item_name_fallback", lambda s: first_nonempty(s, "품목명 없음")),
                    제품구분=("_product_class_fallback", lambda s: first_nonempty(s, "미분류")),
                    생산=("생산", "sum"),
                    계획=("계획", "sum"),
                    잔량=("잔량", "sum"),
                    현재고=("현재고", "max"),
                    재고금액=("재고금액", "max"),
                    진행상태=("_status_fallback", lambda s: first_nonempty(s, "미확인")),
                )
                .reset_index()
            )
            major_items["통합수주건키"] = group_key

        date_values = group["기준확정납기_dt"].dropna()
        if date_values.empty:
            continue
        start_date = date_values.min().date().isoformat()
        end_date = date_values.max().date().isoformat()

        total_plan = float(major_items["계획"].sum())
        total_prod = float(major_items["생산"].sum())
        total_remaining = float(major_items["잔량"].sum())
        total_stock_qty = float(major_items["현재고"].sum())
        total_stock_amount = float(major_items["재고금액"].sum())
        progress_rate = 0 if total_plan <= 0 else total_prod / total_plan
        due_days = (date.fromisoformat(end_date) - today_kst()).days

        if due_days <= 7 and progress_rate < 0.6:
            risk = "높음"
        elif due_days <= 14 and progress_rate < 0.85:
            risk = "보통"
        else:
            risk = "낮음"

        top_item = major_items.sort_values(["수량", "생산"], ascending=[False, False]).iloc[0]
        representative_name = first_nonempty(group["대표프로젝트명"], first_nonempty(group["대표수주건명"], group_key))
        representative_title = first_nonempty(group["상세건명"], representative_name)
        display_name = build_display_name(representative_title, representative_name)
        dealer_source_text = " ".join(
            (
                group.get("_na_dealer", pd.Series(dtype=object)).astype(str).fillna("")
                + " "
                + group.get("_na_perf_dealer", pd.Series(dtype=object)).astype(str).fillna("")
            ).tolist()
        )
        matching_keywords = find_matching_keywords(dealer_source_text, na_active_keywords)
        if "제품구분" in group.columns:
            if product_family == "\uc548\uc131":
                north_america_product_match = group["제품구분"].astype(str).str.contains("\uc548\uc131", na=False).any()
            else:
                north_america_product_match = (
                    group["제품구분"].astype(str).isin(north_america_product_types).any()
                    if north_america_product_types
                    else False
                )
        else:
            north_america_product_match = False
        representative_site = first_nonempty(group["기본주소"], "주소 미등록")
        representative_office = first_nonempty(group["대표사업소"], "사업소 미확인")
        representative_type = "수출" if group["대표사업소"].astype(str).str.contains("수출", na=False).any() else "내수"
        is_north_america = bool(matching_keywords) and north_america_product_match and representative_type == "수출"

        order_records.append(
            {
                "id": group_key,
                "orderNo": first_nonempty(group["수주번호"], group_key),
                "type": representative_type,
                "office": representative_office,
                "customer": representative_name,
                "displayName": display_name,
                "secondaryName": representative_name,
                "site": representative_site,
                "title": representative_title,
                "startDate": start_date,
                "endDate": end_date,
                "items": int(major_items.shape[0]),
                "groupedOrders": int(group["수주번호_norm"].nunique()),
                "amount": int(major_items["수량"].sum()),
                "category": top_item["제품구분"],
                "color": top_item["색상"],
                "code": top_item["단품코드"],
                "risk": risk,
                "plannedQty": int(total_plan),
                "producedQty": int(total_prod),
                "remainingQty": int(total_remaining),
                "stockQty": int(total_stock_qty),
                "stockAmount": int(total_stock_amount),
                "progressRate": progress_rate,
                "isNorthAmerica": is_north_america,
                "northAmericaKeywords": ", ".join(matching_keywords) if is_north_america else "",
                "isEtc": group_key in etc_group_keys and group_key not in major_group_keys,
            }
        )

        items_by_order[group_key] = [
            {
                "name": row["품목명"],
                "spec": row["제품구분"],
                "color": row["색상"],
                "qty": int(row["수량"]),
                "code": row["단품코드"],
                "produced": int(row["생산"]),
                "planned": int(row["계획"]),
                "remaining": int(row["잔량"]),
                "stockQty": int(row["현재고"]),
                "stockAmount": int(row["재고금액"]),
                "status": row["진행상태"],
            }
            for _, row in major_items.sort_values(["수량", "생산"], ascending=[False, False]).iterrows()
        ]

        related_source = group.copy()
        related_source["_기준수량"] = related_source["수주량_num"].where(
            related_source["표준구분"].astype(str).str.contains("주문품", na=False),
            0,
        )
        related_rows_df = (
            related_source.groupby(["수주번호_norm", "수주번호"], dropna=False)
            .agg(
                관련수주건명=("상세건명", lambda s: first_nonempty(s, representative_name)),
                확정납기=("기준확정납기_dt", "max"),
                사업소=("대표사업소", lambda s: first_nonempty(s, representative_office)),
                기준수량=("_기준수량", "sum"),
            )
            .reset_index()
            .sort_values(["확정납기", "수주번호"])
        )
        related_by_order[group_key] = [
            {
                "통합 수주건키": group_key,
                "관련 수주번호": row["수주번호"],
                "관련 수주건명": row["관련수주건명"],
                "확정납기": row["확정납기"].date().isoformat() if pd.notna(row["확정납기"]) else "",
                "사업소": row["사업소"],
                "기준 수량": int(row["기준수량"]),
            }
            for _, row in related_rows_df.iterrows()
        ]

        detail_item_name_col = get_existing_column(group, ["단품명칭", "품목명약칭", "품목명", "품목약칭"])
        detail_standard_col = get_existing_column(group, ["표준구분", "수지구분"])
        detail_brand_col = get_existing_column(group, ["브랜드표시", "브랜드", "브랜드명", "관리브랜드", "관리브랜드명"])
        detail_source = group.copy()
        if detail_item_name_col:
            detail_source["_품목명표시"] = detail_source[detail_item_name_col].astype(str)
        else:
            detail_source["_품목명표시"] = "품목명 없음"
        if detail_standard_col:
            detail_source["_표준구분표시"] = detail_source[detail_standard_col].astype(str)
        else:
            detail_source["_표준구분표시"] = ""
        if detail_brand_col:
            detail_source["_브랜드표시"] = detail_source[detail_brand_col].astype(str)
        else:
            detail_source["_브랜드표시"] = ""
        detail_items_df = (
            detail_source.groupby(["수주번호", "단품코드", "색상"], dropna=False)
            .agg(
                관련수주건명=("상세건명", lambda s: first_nonempty(s, representative_name)),
                브랜드=("_브랜드표시", lambda s: first_nonempty(s, "")),
                품목명=("_품목명표시", lambda s: first_nonempty(s, "품목명 없음")),
                제품구분=("제품구분", lambda s: first_nonempty(s, "미분류")),
                표준구분=("_표준구분표시", lambda s: first_nonempty(s, "")),
                수량=("수주량_num", "sum"),
                현재고=("현재고", "max"),
                예량D7=("예량D7", "min"),
                # 재고/예측 값은 수주라인 중복으로 증폭되지 않도록 대표값(max) 사용
                출고예정D7=("출고예정D7", "max"),
                확정납기=("기준확정납기_dt", "max"),
                사업소=("대표사업소", lambda s: first_nonempty(s, representative_office)),
            )
            .reset_index()
            .sort_values(["확정납기", "수주번호", "품목명"])
        )
        detail_items_by_order[group_key] = [
            {
                "관련 수주번호": row["수주번호"],
                "관련 수주건명": row["관련수주건명"],
                "브랜드": row["브랜드"],
                "품목명": row["품목명"],
                "제품구분": row["제품구분"],
                "표준구분": row["표준구분"],
                "단품코드": row["단품코드"],
                "색상": row["색상"],
                "수량": int(row["수량"]),
                "현재고": int(row["현재고"]) if pd.notna(row["현재고"]) else 0,
                "예량(D+7)": int(row["예량D7"]) if pd.notna(row["예량D7"]) else 0,
                "출고예정(D+7)": int(row["출고예정D7"]) if pd.notna(row["출고예정D7"]) else 0,
                "확정납기": row["확정납기"].date().isoformat() if pd.notna(row["확정납기"]) else "",
                "사업소": row["사업소"],
            }
            for _, row in detail_items_df.iterrows()
        ]
    step_start = record_perf_step(perf_rows, "order_record_build", step_start)

    order_records.sort(key=lambda item: (item["startDate"], item["displayName"]))
    available_months_set = set()
    for record in order_records:
        try:
            start_day = date.fromisoformat(str(record.get("startDate", "")))
            end_day = date.fromisoformat(str(record.get("endDate", "")))
        except Exception:
            continue
        if end_day < start_day:
            start_day, end_day = end_day, start_day
        cursor = date(start_day.year, start_day.month, 1)
        end_month = date(end_day.year, end_day.month, 1)
        while cursor <= end_month:
            available_months_set.add(f"{cursor.year:04d}-{cursor.month:02d}")
            if cursor.month == 12:
                cursor = date(cursor.year + 1, 1, 1)
            else:
                cursor = date(cursor.year, cursor.month + 1, 1)
    available_months = sorted(available_months_set)
    step_start = record_perf_step(perf_rows, "available_months_build", step_start)

    return {
        "orders": order_records,
        "items_by_order": items_by_order,
        "related_by_order": related_by_order,
        "detail_items_by_order": detail_items_by_order,
        "available_months": available_months,
        "source_titles": latest_titles,
        "north_america_keyword_rows": na_keyword_rows,
        "north_america_active_keywords": na_active_keywords,
        "perf_load_ms": perf_rows,
    }


def get_month_days(year: int, month: int):
    first_day = date(year, month, 1)
    _, total_days = calendar.monthrange(year, month)
    start_weekday = (first_day.weekday() + 1) % 7
    cells = [None] * start_weekday
    for day in range(1, total_days + 1):
        cells.append(date(year, month, day))
    while len(cells) % 7 != 0:
        cells.append(None)
    return cells


def style_badge(label: str, palette: dict[str, str]) -> str:
    return (
        f"<span class='badge' style='background:{palette['bg']};"
        f"color:{palette['text']};border-color:{palette['border']};'>{label}</span>"
    )


def initialize_state():
    if "business_type" not in st.session_state:
        st.session_state["business_type"] = "전체"
    if "view_style" not in st.session_state:
        st.session_state["view_style"] = "구성요소"
    if "selected_month" not in st.session_state:
        st.session_state["selected_month"] = "2026-03"
    if "selected_order_id" not in st.session_state:
        st.session_state["selected_order_id"] = "ORD-001"
    if "selected_order_ids" not in st.session_state:
        st.session_state["selected_order_ids"] = ["ORD-001"]
    if "detail_selected_order_id" not in st.session_state:
        st.session_state["detail_selected_order_id"] = "ORD-001"
    if "detail_metric" not in st.session_state:
        st.session_state["detail_metric"] = None
    if "detail_order_ids" not in st.session_state:
        st.session_state["detail_order_ids"] = []
    if "detail_metric_open_ts" not in st.session_state:
        st.session_state["detail_metric_open_ts"] = 0
    if "drilldown_order_id" not in st.session_state:
        st.session_state["drilldown_order_id"] = None
    if "last_summary_click_ts" not in st.session_state:
        st.session_state["last_summary_click_ts"] = 0
    if "last_calendar_click_ts" not in st.session_state:
        st.session_state["last_calendar_click_ts"] = 0
    if "etc_amount_threshold" not in st.session_state:
        st.session_state["etc_amount_threshold"] = 100_000_000
    if "etc_amount_threshold_input" not in st.session_state:
        st.session_state["etc_amount_threshold_input"] = 100_000_000
    if "etc_amount_threshold_input_text" not in st.session_state:
        st.session_state["etc_amount_threshold_input_text"] = "100,000,000"
    if "product_family" not in st.session_state:
        st.session_state["product_family"] = "\ucda9\uc8fc"
    if "export_only_nonstock_custom" not in st.session_state:
        st.session_state["export_only_nonstock_custom"] = False
    if "export_filter_mode" not in st.session_state:
        st.session_state["export_filter_mode"] = "비재고/주문품" if st.session_state["export_only_nonstock_custom"] else "전체"


def get_filtered_orders(data: dict):
    month_value = st.session_state["selected_month"]
    business_type = st.session_state["business_type"]
    year, month = map(int, month_value.split("-"))
    month_start = date(year, month, 1)
    month_end = date(year, month, calendar.monthrange(year, month)[1])
    filtered = []
    for order in data["orders"]:
        try:
            order_start = date.fromisoformat(str(order.get("startDate", "")))
            order_end = date.fromisoformat(str(order.get("endDate", "")))
            if order_end < order_start:
                order_start, order_end = order_end, order_start
            month_match = not (order_end < month_start or order_start > month_end)
        except Exception:
            month_match = False
        type_match = business_type == "전체" or order["type"] == business_type
        if month_match and type_match:
            filtered.append(order)
    return filtered


def day_orders(current_day: date, filtered_orders: list[dict]):
    if current_day is None:
        return []
    current_key = current_day.isoformat()
    return [
        order
        for order in filtered_orders
        if order["startDate"] <= current_key <= order["endDate"]
    ]


def build_overlay_calendar_payload(filtered_orders: list[dict], selected_month: str, view_style: str):
    year, month = map(int, selected_month.split("-"))
    month_calendar = calendar.Calendar(firstweekday=6).monthdatescalendar(year, month)
    today = today_kst()
    weeks = []

    for week_days in month_calendar:
        week_start = week_days[0]
        week_end = week_days[-1]
        week_events = []

        for order in filtered_orders:
            order_start = date.fromisoformat(order["startDate"])
            order_end = date.fromisoformat(order["endDate"])
            if order_end < week_start or order_start > week_end:
                continue

            visible_start = max(order_start, week_start)
            visible_end = min(order_end, week_end)
            start_col = max(0, (visible_start - week_start).days)
            end_col = min(6, (visible_end - week_start).days)

            if visible_start == order_start and visible_end == order_end:
                shape = "single"
            elif visible_start == order_start:
                shape = "start"
            elif visible_end == order_end:
                shape = "end"
            else:
                shape = "middle"

            label = order["displayName"] if view_style == "구성요소" else order["title"]
            week_events.append(
                {
                    "group_key": order["id"],
                    "label": clip_text(label, 22),
                    "title": f"{order['displayName']} | {order['startDate']} ~ {order['endDate']}",
                    "start_col": start_col,
                    "end_col": end_col,
                    "color_type": "domestic" if order["type"] == "내수" else "export",
                    "shape": shape,
                }
            )

        week_events.sort(key=lambda item: (item["start_col"], item["end_col"], item["label"]))
        lanes: list[list[dict]] = []
        for event in week_events:
            placed = False
            for lane in lanes:
                if all(event["end_col"] < existing["start_col"] or event["start_col"] > existing["end_col"] for existing in lane):
                    lane.append(event)
                    placed = True
                    break
            if not placed:
                lanes.append([event])

        weeks.append(
            {
                "week_key": f"{week_days[0].isoformat()}_{week_days[-1].isoformat()}",
                "days": [
                    {
                        "day": day.day,
                        "in_month": day.month == month,
                        "is_today": day == today,
                        "weekday_index": idx,
                    }
                    for idx, day in enumerate(week_days)
                ],
                "lanes": lanes,
            }
        )

    return {"weekday_labels": WEEKDAY_LABELS, "weeks": weeks}


def sync_selected_order_from_detail():
    st.session_state["selected_order_id"] = st.session_state["detail_selected_order_id"]


def reset_detail_views():
    st.session_state["detail_metric"] = None
    st.session_state["drilldown_order_id"] = None
    st.session_state["detail_order_ids"] = []
    st.session_state["detail_metric_open_ts"] = 0


def on_top_filter_change():
    # Changing top filters should close stale popups from previous context.
    reset_detail_views()


def render_header(data: dict):
    snapshot_label = build_source_snapshot_label(data.get("source_titles", {}))
    st.markdown(
        f"""
        <div class="hero-card">
            <div style="display:flex; align-items:flex-end; justify-content:space-between; gap:16px;">
                <div style="font-size:34px; font-weight:800; color:#0f172a;">품평/프로젝트 관리</div>
                <div style="font-size:13px; color:#475569; font-weight:600;">
                    기준: {snapshot_label}
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_north_america_keyword_manager(data: dict):
    source_rows = data.get("north_america_keyword_rows", []) or [
        {"키워드": kw, "활성": True} for kw in DEFAULT_NORTH_AMERICA_KEYWORDS
    ]
    source_signature = "|".join(f"{row.get('키워드','')}::{int(bool(row.get('활성', True)))}" for row in source_rows)
    if st.session_state.get("na_keyword_source_signature") != source_signature:
        st.session_state["na_keyword_source_signature"] = source_signature
        st.session_state["na_keyword_editor_df"] = pd.DataFrame(source_rows)

    with st.expander("북미 키워드 관리", expanded=False):
        st.caption("수주건명에 아래 키워드가 포함되면 북미 건으로 표시됩니다.")
        edited_df = st.data_editor(
            st.session_state["na_keyword_editor_df"],
            key="na_keyword_editor_widget",
            num_rows="dynamic",
            use_container_width=True,
            hide_index=True,
        )
        if st.button("북미 키워드 저장", key="save_na_keywords_btn", use_container_width=True):
            rows = edited_df.to_dict("records")
            save_north_america_keywords(rows)
            load_dashboard_data.clear()
            st.success("북미 키워드를 저장했습니다. 대시보드를 새로 불러옵니다.")
            st.rerun()


def metric_rows(filtered_orders: list[dict]):
    return {
        "totalOrders": [
            {"name": order["displayName"], "sub": order["secondaryName"], "value": f"{order['startDate']} ~ {order['endDate']}", "order": order}
            for order in filtered_orders
        ],
        "groupedCount": [
            {"name": order["displayName"], "sub": order["office"], "value": f"{order['groupedOrders']}", "order": order}
            for order in filtered_orders
        ],
        "itemCount": [
            {"name": f"{order['category']} / {order['code']}", "sub": f"색상 {order['color']}", "value": f"{order['items']}", "order": order}
            for order in filtered_orders
        ],
        "totalAmount": [
            {"name": order["displayName"], "sub": order["category"], "value": f"{order['amount']:,}", "order": order}
            for order in filtered_orders
        ],
    }


def render_calendar_and_detail(filtered_orders: list[dict], data: dict, available_months: list[str]):
    order_by_id = {order["id"]: order for order in filtered_orders}
    valid_ids = set(order_by_id.keys())
    selected_ids = st.session_state.get("selected_order_ids", [])
    if not isinstance(selected_ids, list):
        selected_ids = []
    selected_ids = [order_id for order_id in selected_ids if order_id in valid_ids]

    selected_id = st.session_state.get("selected_order_id", "")
    if selected_id in valid_ids and selected_id not in selected_ids:
        selected_ids = [selected_id]
    if not selected_ids and filtered_orders:
        selected_ids = [filtered_orders[0]["id"]]

    st.session_state["selected_order_ids"] = selected_ids
    if selected_ids and st.session_state.get("selected_order_id") not in selected_ids:
        st.session_state["selected_order_id"] = selected_ids[0]
    if st.session_state.get("detail_selected_order_id") != st.session_state.get("selected_order_id"):
        st.session_state["detail_selected_order_id"] = st.session_state.get("selected_order_id")

    left_col, right_col = st.columns([1, 1])
    with left_col:
        with st.container(border=True):
            st.radio(
                "\uc0ac\uc5c5\uc7a5",
                options=["\ucda9\uc8fc", "\uc548\uc131"],
                horizontal=True,
                key="product_family",
                on_change=on_top_filter_change,
            )
            filter_col1, filter_col2 = st.columns([1.35, 1])
            with filter_col1:
                st.radio("사업소 필터", ["전체", "내수", "수출"], horizontal=True, key="business_type", on_change=on_top_filter_change)
            with filter_col2:
                st.selectbox("조회 월", available_months, format_func=month_label, key="selected_month", on_change=on_top_filter_change)
            payload = build_overlay_calendar_payload(filtered_orders, st.session_state["selected_month"], st.session_state["view_style"])
            clicked_result = overlay_calendar_component(
                data=payload,
                selected_group_key=st.session_state["selected_order_id"],
                selected_group_keys=st.session_state.get("selected_order_ids", []),
                key="overlay_calendar",
                default=None,
            )
            clicked_order_id = None
            selected_keys_from_component = None
            if isinstance(clicked_result, dict):
                clicked_order_id = clicked_result.get("lastClicked")
                raw_keys = clicked_result.get("selectedKeys")
                if isinstance(raw_keys, list):
                    selected_keys_from_component = [
                        str(order_id) for order_id in raw_keys if str(order_id) in valid_ids
                    ]
            elif isinstance(clicked_result, str):
                clicked_order_id = clicked_result
                if clicked_order_id in valid_ids:
                    selected_keys_from_component = [clicked_order_id]

            changed = False
            calendar_interacted = clicked_result is not None
            if selected_keys_from_component is not None:
                if selected_keys_from_component != st.session_state.get("selected_order_ids", []):
                    st.session_state["selected_order_ids"] = selected_keys_from_component
                    changed = True
            if clicked_order_id and clicked_order_id in valid_ids and clicked_order_id != st.session_state["selected_order_id"]:
                st.session_state["selected_order_id"] = clicked_order_id
                st.session_state["detail_selected_order_id"] = clicked_order_id
                if clicked_order_id not in st.session_state.get("selected_order_ids", []):
                    st.session_state["selected_order_ids"] = [clicked_order_id]
                changed = True
            if calendar_interacted and (
                st.session_state.get("detail_metric") is not None
                or st.session_state.get("drilldown_order_id") is not None
                or bool(st.session_state.get("detail_order_ids"))
            ):
                # 캘린더 클릭 시 요약/드릴다운 팝업 상태는 닫는다.
                st.session_state["detail_metric"] = None
                st.session_state["detail_order_ids"] = []
                st.session_state["drilldown_order_id"] = None
                changed = True
            if changed:
                st.rerun()

    with right_col:
        with st.container(border=True):
            st.selectbox(
                "상세 수주건 선택",
                options=[order["id"] for order in filtered_orders],
                key="detail_selected_order_id",
                on_change=sync_selected_order_from_detail,
                format_func=lambda order_id: next(
                    (f"{o['startDate']} ~ {o['endDate']} | {o['displayName']}" for o in filtered_orders if o["id"] == order_id),
                    order_id,
                ),
            )
            selected_order = next((order for order in filtered_orders if order["id"] == st.session_state["selected_order_id"]), None)
            if not selected_order:
                st.info("선택 가능한 수주건이 없습니다.")
                return

            c1, c2 = st.columns(2)
            c1.metric("시작 보고기", format_korean_date(selected_order["startDate"]))
            c2.metric("종료 보고기", format_korean_date(selected_order["endDate"]))

            c3, c4 = st.columns(2)
            c3.metric("포함 수주건 수", f"{selected_order['items']}")
            na_badge = style_badge("북미", NA_COLORS) if selected_order.get("isNorthAmerica") else ""
            c4.markdown(
                style_badge(selected_order["type"], TYPE_COLORS[selected_order["type"]])
                + style_badge(f"리스크 {selected_order['risk']}", RISK_COLORS[selected_order["risk"]])
                + na_badge,
                unsafe_allow_html=True,
            )

            p1, p2, p3 = st.columns(3)
            p1.metric("생산 계획", f"{selected_order['plannedQty']:,}")
            p2.metric("생산 실적", f"{selected_order['producedQty']:,}")
            p3.metric("잔량", f"{selected_order['remainingQty']:,}")

            s1, s2 = st.columns(2)
            s1.metric("현재고", f"{selected_order.get('stockQty', 0):,}")
            s2.metric("재고금액", f"{selected_order.get('stockAmount', 0):,}")

            related_rows = data["related_by_order"].get(selected_order["id"], [])
            st.dataframe(
                format_df_numbers_for_display(with_row_no_and_total(pd.DataFrame(related_rows))),
                use_container_width=True,
                hide_index=True,
            )


def render_metrics(filtered_orders: list[dict]):
    selected_month = st.session_state["selected_month"]
    year, month = map(int, selected_month.split("-"))
    today = today_kst()
    if today.year == year and today.month == month:
        ref_date = today
    else:
        ref_date = date(year, month, 1)

    week_no = ((ref_date.day - 1) // 7) + 1
    week_start = ref_date - timedelta(days=ref_date.weekday())
    week_end = week_start + timedelta(days=6)

    def overlaps_week(order: dict) -> bool:
        start = date.fromisoformat(order["startDate"])
        end = date.fromisoformat(order["endDate"])
        return not (end < week_start or start > week_end)

    weekly_orders = [order for order in filtered_orders if overlaps_week(order)]
    st.session_state["weekly_order_ids"] = [order["id"] for order in weekly_orders]
    st.session_state["monthly_order_ids"] = [order["id"] for order in filtered_orders]
    weekly_export = sum(1 for order in weekly_orders if order["type"] == "수출")
    weekly_domestic = sum(1 for order in weekly_orders if order["type"] == "내수")
    monthly_export = sum(1 for order in filtered_orders if order["type"] == "수출")
    monthly_domestic = sum(1 for order in filtered_orders if order["type"] == "내수")

    title_week = f"{str(year)[-2:]}년 {month}월 {week_no}주 주요 수주건"
    title_month = f"{str(year)[-2:]}년 {month}월 주요 수주건"

    card_payload = {
        "weekly": {
            "subtitle": title_week,
            "total": f"{len(weekly_orders):,}",
            "partial": f"수출 {weekly_export:,}건 / 내수 {weekly_domestic:,}건",
        },
        "monthly": {
            "subtitle": title_month,
            "total": f"{len(filtered_orders):,}",
            "partial": f"수출 {monthly_export:,}건 / 내수 {monthly_domestic:,}건",
        },
    }

    clicked_summary = summary_cards_component(
        data=card_payload,
        key="summary_cards_click",
        default=None,
    )
    if clicked_summary == "weekly":
        st.session_state["detail_metric"] = "totalOrders"
        st.session_state["detail_order_ids"] = st.session_state.get("weekly_order_ids", [])
    elif clicked_summary == "monthly":
        st.session_state["detail_metric"] = "groupedCount"
        st.session_state["detail_order_ids"] = st.session_state.get("monthly_order_ids", [])


@st.dialog("세부내역 보기", width="large")
def show_metric_detail_dialog(data: dict, filtered_orders: list[dict]):
    detail_metric = st.session_state.get("detail_metric")
    if not detail_metric:
        st.info("표시할 세부내역이 없습니다.")
        return

    selected_ids = list(dict.fromkeys(st.session_state.get("detail_order_ids", [])))
    order_pool = filtered_orders if filtered_orders else data.get("orders", [])
    order_map = {order["id"]: order for order in order_pool}
    if selected_ids:
        source_orders = [order_map[order_id] for order_id in selected_ids if order_id in order_map]
    else:
        source_orders = order_pool

    st.markdown("### 수주건별 세부 품목 리스트")
    st.caption("상단 요약 현황에서 선택된 수주건 기준입니다.")

    info_cols = st.columns(3)
    info_cols[0].markdown(
        f"<div class='soft-card'><div class='subtle-title'>조회 월</div><div style='font-size:20px;font-weight:700;color:#0f172a;'>{month_label(st.session_state['selected_month'])}</div></div>",
        unsafe_allow_html=True,
    )
    info_cols[1].markdown(
        f"<div class='soft-card'><div class='subtle-title'>사업 구분</div><div style='font-size:20px;font-weight:700;color:#0f172a;'>{st.session_state['business_type']}</div></div>",
        unsafe_allow_html=True,
    )
    info_cols[2].markdown(
        f"<div class='soft-card'><div class='subtle-title'>선택 수주건</div><div style='font-size:20px;font-weight:700;color:#0f172a;'>{len(source_orders):,}건</div></div>",
        unsafe_allow_html=True,
    )

    if st.session_state.get("summary_popup_product_filter") == "목제":
        st.session_state["summary_popup_product_filter"] = "목제상품"

    filter_cols = st.columns(3)
    standard_filter = filter_cols[0].selectbox(
        "표준구분",
        options=["전체", "표준품", "주문품"],
        index=2,
        key="summary_popup_standard_filter",
    )
    product_filter = filter_cols[1].selectbox(
        "제품구분",
        options=["전체", "충주", "목제상품"],
        index=1,
        key="summary_popup_product_filter",
    )
    return_only_filter = filter_cols[2].toggle(
        "회수",
        key="summary_popup_return_filter",
    )

    chgju_products = {"충주1제품", "충주2제품", "F우레탄제품"}
    wood_products = {"F우레탄제품", "베트남상품", "목제상품", "목제5상품", "목제6상품"}

    merged_item_rows = []
    for order in source_orders:
        for row in data.get("detail_items_by_order", {}).get(order["id"], []):
            row_copy = dict(row)
            row_copy["대표 수주건명"] = order.get("displayName", "")
            merged_item_rows.append(row_copy)

    display_df = pd.DataFrame(merged_item_rows)
    if not display_df.empty:
        if standard_filter in {"주문품", "표준품"} and "표준구분" in display_df.columns:
            display_df = display_df[display_df["표준구분"].astype(str).str.contains(standard_filter, na=False)]
        if product_filter != "전체" and "제품구분" in display_df.columns:
            if product_filter == "충주":
                allowed_products = chgju_products
            else:
                allowed_products = chgju_products | wood_products
            display_df = display_df[display_df["제품구분"].isin(allowed_products)]

        display_df = display_df.rename(columns={"품목명": "단품명칭", "수량": "수주량"})
        if return_only_filter and "현재고" in display_df.columns and "수주량" in display_df.columns:
            stock_series = pd.to_numeric(display_df["현재고"], errors="coerce").fillna(0)
            order_qty_series = pd.to_numeric(display_df["수주량"], errors="coerce").fillna(0)
            display_df = display_df[stock_series > order_qty_series]
        if "단품명칭" in display_df.columns:
            display_df["단품명칭"] = display_df["단품명칭"].map(shorten_item_name_for_display)

    wanted_cols = [
        "대표 수주건명",
        "브랜드",
        "제품구분",
        "단품코드",
        "색상",
        "단품명칭",
        "수주량",
        "현재고",
        "확정납기",
        "예량(D+7)",
        "출고예정(D+7)",
    ]
    display_df = display_df[[col for col in wanted_cols if col in display_df.columns]] if not display_df.empty else display_df

    if not display_df.empty:
        if {"단품코드", "색상"}.issubset(display_df.columns):
            popup_total_items = int(
                display_df[["단품코드", "색상"]].astype(str).drop_duplicates().shape[0]
            )
        else:
            popup_total_items = int(len(display_df))
        popup_total_amount = int(
            pd.to_numeric(display_df.get("수주량", pd.Series(dtype=float)), errors="coerce").fillna(0).sum()
        )
    else:
        popup_total_items = 0
        popup_total_amount = 0

    sum_cols = st.columns(2)
    sum_cols[0].markdown(
        f"<div class='soft-card'><div class='subtle-title'>합계 품목수</div><div style='font-size:20px;font-weight:700;color:#0f172a;'>{popup_total_items:,} 품목</div></div>",
        unsafe_allow_html=True,
    )
    sum_cols[1].markdown(
        f"<div class='soft-card'><div class='subtle-title'>합계 수주량</div><div style='font-size:20px;font-weight:700;color:#0f172a;'>{popup_total_amount:,} BOX</div></div>",
        unsafe_allow_html=True,
    )

    if display_df.empty:
        st.info("현재 필터 기준으로 표시할 품목이 없습니다.")
    else:
        display_df_view = with_row_no_and_total(display_df)
        st.dataframe(format_df_numbers_for_display(display_df_view), use_container_width=True, hide_index=True)
    action_cols = st.columns(2)
    with action_cols[0]:
        if display_df.empty:
            st.button(
                "다운로드 (.xlsx)",
                key="summary_popup_download_excel_disabled",
                disabled=True,
                use_container_width=True,
            )
        else:
            excel_bytes = dataframe_to_styled_excel_bytes(display_df_view, sheet_name="통합품목")
            st.download_button(
                "다운로드 (.xlsx)",
                data=excel_bytes,
                file_name=f"통합품목_요약팝업_{today_kst().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="summary_popup_download_excel",
            )
    with action_cols[1]:
        close_clicked = st.button("닫기", key="close_metric_detail", use_container_width=True)

    if close_clicked:
        st.session_state["detail_metric"] = None
        st.session_state["detail_order_ids"] = []
        st.session_state["detail_metric_open_ts"] = 0
        st.rerun()


@st.dialog("품목 리스트", width="large")
def show_drilldown_dialog(data: dict):
    drill_id = st.session_state.get("drilldown_order_id")
    drill_order = next((order for order in data["orders"] if order["id"] == drill_id), None)
    if not drill_order:
        st.info("표시할 품목 정보가 없습니다.")
        return

    st.markdown(f"### {drill_order['displayName']} 품목 리스트")
    st.caption("선택한 수주건의 품목 목록입니다.")

    info_cols = st.columns(3)
    info_cols[0].markdown(f"<div class='soft-card'><div class='subtle-title'>수주건명</div><div style='font-size:20px;font-weight:700;color:#0f172a;'>{drill_order['displayName']}</div></div>", unsafe_allow_html=True)
    info_cols[1].markdown(f"<div class='soft-card'><div class='subtle-title'>대표명</div><div style='font-size:20px;font-weight:700;color:#0f172a;'>{drill_order['secondaryName']}</div></div>", unsafe_allow_html=True)
    info_cols[2].markdown(f"<div class='soft-card'><div class='subtle-title'>품목 수</div><div style='font-size:20px;font-weight:700;color:#0f172a;'>{len(data['items_by_order'].get(drill_id, []))}</div></div>", unsafe_allow_html=True)

    item_df = pd.DataFrame(data["items_by_order"].get(drill_id, []))
    item_df = item_df.rename(columns={"name": "품목명", "spec": "규격", "color": "색상", "qty": "수량", "code": "단품코드", "produced": "생산", "planned": "계획", "remaining": "잔량", "stockQty": "현재고", "stockAmount": "재고금액", "status": "진행상태"})
    st.dataframe(
        format_df_numbers_for_display(with_row_no_and_total(item_df)),
        use_container_width=True,
        hide_index=True,
    )

    if st.button("닫기", key="close_drilldown_detail", use_container_width=True):
        st.session_state["drilldown_order_id"] = None
        st.rerun()


def render_dialogs(filtered_orders: list[dict], data: dict):
    # Streamlit allows only one dialog per run.
    # Give drilldown dialog priority, otherwise show metric dialog.
    if st.session_state.get("drilldown_order_id"):
        st.session_state["detail_metric"] = None
        show_drilldown_dialog(data)
    elif st.session_state.get("detail_metric"):
        show_metric_detail_dialog(data, filtered_orders)


def render_calendar_and_detail(filtered_orders: list[dict], data: dict, available_months: list[str]):
    order_by_id = {order["id"]: order for order in filtered_orders}
    valid_ids = set(order_by_id.keys())
    selected_ids = st.session_state.get("selected_order_ids", [])
    if not isinstance(selected_ids, list):
        selected_ids = []
    selected_ids = [order_id for order_id in selected_ids if order_id in valid_ids]

    selected_id = st.session_state.get("selected_order_id", "")
    if selected_id in valid_ids and selected_id not in selected_ids:
        selected_ids = [selected_id]
    if not selected_ids and filtered_orders:
        selected_ids = [filtered_orders[0]["id"]]

    st.session_state["selected_order_ids"] = selected_ids
    if selected_ids and st.session_state.get("selected_order_id") not in selected_ids:
        st.session_state["selected_order_id"] = selected_ids[0]
    if st.session_state.get("detail_selected_order_id") != st.session_state.get("selected_order_id"):
        st.session_state["detail_selected_order_id"] = st.session_state.get("selected_order_id")

    left_col, right_col = st.columns([1, 1])
    with left_col:
        with st.container(border=True):
            st.radio(
                "\uc0ac\uc5c5\uc7a5",
                options=["충주", "안성"],
                horizontal=True,
                key="product_family",
                on_change=on_top_filter_change,
            )
            filter_col1, filter_col2 = st.columns([1.35, 1])
            with filter_col1:
                st.radio(
                    "사업소 필터",
                    options=["전체", "내수", "수출"],
                    horizontal=True,
                    key="business_type",
                    on_change=on_top_filter_change,
                    label_visibility="visible",
                )
            with filter_col2:
                st.selectbox(
                    "조회 월",
                    options=available_months,
                    format_func=month_label,
                    key="selected_month",
                    on_change=on_top_filter_change,
                    label_visibility="visible",
                )
            # 북미 기준은 스프레드시트 시트(북미 딜러관리)에서 직접 관리
            payload = build_overlay_calendar_payload(
                filtered_orders,
                st.session_state["selected_month"],
                st.session_state["view_style"],
            )
            clicked_order_id = overlay_calendar_component(
                data=payload,
                selected_group_key=st.session_state["selected_order_id"],
                key="overlay_calendar",
                default=None,
            )
            if clicked_order_id and clicked_order_id != st.session_state["selected_order_id"]:
                st.session_state["selected_order_id"] = clicked_order_id
                st.session_state["detail_selected_order_id"] = clicked_order_id
                selected_order = next((order for order in filtered_orders if order["id"] == clicked_order_id), selected_order)
                st.rerun()

    with right_col:
        with st.container(border=True):
            st.selectbox(
                "상세 수주건 선택",
                options=[order["id"] for order in filtered_orders],
                key="detail_selected_order_id",
                on_change=sync_selected_order_from_detail,
                format_func=lambda order_id: next(
                    (
                        f"{order['startDate']} ~ {order['endDate']} | {order['displayName']}"
                        for order in filtered_orders
                        if order["id"] == order_id
                    ),
                    order_id,
                ),
            )

            selected_order = next((order for order in filtered_orders if order["id"] == st.session_state["selected_order_id"]), None)
            if not selected_order:
                st.info("선택 가능한 수주건이 없습니다.")
                return

            metric_cols = st.columns(2)
            metric_cols[0].metric("시작 보고기", format_korean_date(selected_order["startDate"]))
            metric_cols[1].metric("종료 보고기", format_korean_date(selected_order["endDate"]))

            metric_cols2 = st.columns(2)
            metric_cols2[0].metric("포함 수주건 수", f"{selected_order['items']}")
            na_badge = style_badge("북미", NA_COLORS) if selected_order.get("isNorthAmerica") else ""
            metric_cols2[1].markdown(
                style_badge(selected_order["type"], TYPE_COLORS[selected_order["type"]]) +
                style_badge(f"리스크 {selected_order['risk']}", RISK_COLORS[selected_order["risk"]]) +
                na_badge,
                unsafe_allow_html=True,
            )

            production_cols = st.columns(3)
            production_cols[0].metric("생산 계획", f"{selected_order['plannedQty']:,}")
            production_cols[1].metric("생산 실적", f"{selected_order['producedQty']:,}")
            production_cols[2].metric("잔량", f"{selected_order['remainingQty']:,}")

            stock_cols = st.columns(2)
            stock_cols[0].metric("현재고", f"{selected_order.get('stockQty', 0):,}")
            stock_cols[1].metric("재고금액", f"{selected_order.get('stockAmount', 0):,}")

            st.markdown(
                f"""
                <div class="detail-box">
                    <div style="margin-bottom:10px;"><strong>수주건명 :</strong> {selected_order['displayName']}</div>
                    <div style="margin-bottom:10px;"><strong>대표명 :</strong> {selected_order['secondaryName']}</div>
                    <div style="margin-bottom:10px;"><strong>처리 위치 :</strong> {selected_order['site']}</div>
                    <div style="margin-bottom:10px;"><strong>주요 품목 :</strong> {selected_order['category']} / 색상 {selected_order['color']} / 코드 {selected_order['code']}</div>
                    <div><strong>생산 진행률 :</strong> {selected_order['progressRate'] * 100:.1f}%</div>
                </div>
                """,
                unsafe_allow_html=True,
            )

            st.markdown("### 동일한 통합 수주건으로 묶인 관련 정보")
            related_rows = data["related_by_order"].get(selected_order["id"], [])
            st.dataframe(
                format_df_numbers_for_display(with_row_no_and_total(pd.DataFrame(related_rows))),
                use_container_width=True,
                hide_index=True,
            )

            action_cols = st.columns(3)
            action_cols[0].button("공유", key="share_btn", use_container_width=True)
            action_cols[1].button("엑셀 내보내기", key="export_btn", use_container_width=True)
            action_cols[2].button("상세 보고서", key="report_btn", use_container_width=True)


def render_order_list(filtered_orders: list[dict]):
    with st.container(border=True):
        st.markdown("## 수주건 리스트")
        cols = st.columns(3)
        for idx, order in enumerate(filtered_orders):
            with cols[idx % 3]:
                is_selected = st.session_state["selected_order_id"] == order["id"]
                with st.container(border=True):
                    if is_selected:
                        st.markdown(
                            "<div style='margin-bottom:10px; font-size:12px; font-weight:700; color:#0f172a;'>현재 선택된 수주건</div>",
                            unsafe_allow_html=True,
                        )
                    else:
                        st.markdown("<div style='height:26px;'></div>", unsafe_allow_html=True)

                    st.markdown(
                        style_badge(order["type"], TYPE_COLORS[order["type"]]) +
                        style_badge(order["risk"], RISK_COLORS[order["risk"]]),
                        unsafe_allow_html=True,
                    )
                    st.markdown(f"### {order['displayName']}")
                    st.caption(order["secondaryName"])
                    st.markdown(
                        "<div class='tiny'>세부내역 버튼으로 상단 상세 정보를 열 수 있습니다.</div>",
                        unsafe_allow_html=True,
                    )
                    st.write(f"일정: {order['startDate']} ~ {order['endDate']}")
                    st.write(f"사업소: {order['office']}")
                    st.write(f"통합 수주건 수: {order['groupedOrders']}")
                    st.write(f"현장: {order['site']}")
                    if st.button("이 수주건 보기", key=f"order_pick_{order['id']}", use_container_width=True):
                        st.session_state["selected_order_id"] = order["id"]
                        st.rerun()
                    if st.button("세부내역 열기", key=f"order_detail_{order['id']}", use_container_width=True):
                        st.session_state["selected_order_id"] = order["id"]
                        st.session_state["drilldown_order_id"] = order["id"]
                        st.rerun()


NORTH_AMERICA_DEALER_SHEET = "\ubd81\ubbf8 \ub51c\ub7ec\uad00\ub9ac"
NORTH_AMERICA_LEGACY_SHEET = "\ubd81\ubbf8\ud0a4\uc6cc\ub4dc"
NA_COL_KEYWORD = "\ud0a4\uc6cc\ub4dc"
NA_COL_ACTIVE = "\ud65c\uc131"


def normalize_match_text(value: str) -> str:
    text = str(value or "").upper()
    return re.sub(r"[^A-Z0-9]", "", text)


def get_existing_column(df: pd.DataFrame, candidates: list[str]) -> str | None:
    normalized = {re.sub(r"\s+", "", str(col)): col for col in df.columns}
    for candidate in candidates:
        key = re.sub(r"\s+", "", candidate)
        if key in normalized:
            return normalized[key]
    return None


def find_matching_keywords(text: str, keywords: list[str]) -> list[str]:
    normalized_text = normalize_match_text(text)
    if not normalized_text:
        return []
    matches: list[str] = []
    for raw_kw in keywords:
        kw = str(raw_kw or "").strip()
        if not kw:
            continue
        normalized_kw = normalize_match_text(kw)
        if normalized_kw and normalized_kw in normalized_text:
            matches.append(kw)
    return matches


def ensure_north_america_keyword_worksheet(spreadsheet):
    for sheet_name in [NORTH_AMERICA_DEALER_SHEET, NORTH_AMERICA_LEGACY_SHEET]:
        try:
            return spreadsheet.worksheet(sheet_name)
        except WorksheetNotFound:
            continue
    worksheet = spreadsheet.add_worksheet(title=NORTH_AMERICA_DEALER_SHEET, rows=400, cols=2)
    initial_rows = [[NA_COL_KEYWORD, NA_COL_ACTIVE]] + [[kw, "Y"] for kw in DEFAULT_NORTH_AMERICA_KEYWORDS]
    worksheet.update(range_name="A1", values=initial_rows)
    return worksheet


def load_north_america_keywords(spreadsheet) -> tuple[list[dict], list[str]]:
    worksheet = ensure_north_america_keyword_worksheet(spreadsheet)
    values = worksheet.get_all_values()
    if not values:
        rows = [{NA_COL_KEYWORD: kw, NA_COL_ACTIVE: True} for kw in DEFAULT_NORTH_AMERICA_KEYWORDS]
        return rows, DEFAULT_NORTH_AMERICA_KEYWORDS.copy()

    header = [re.sub(r"\s+", "", str(v).strip()) for v in values[0]]
    data_rows = values[1:]

    keyword_candidates = [NA_COL_KEYWORD, "\ub300\ub9ac\uc810", "\uc2e4\uc801\ub300\ub9ac\uc810", "\ucf54\ub4dc", "keyword", "code"]
    active_candidates = [NA_COL_ACTIVE, "active", "use", "enabled"]

    idx_kw = 0
    for cand in keyword_candidates:
        cand_norm = re.sub(r"\s+", "", cand)
        if cand_norm in header:
            idx_kw = header.index(cand_norm)
            break

    idx_on = None
    for cand in active_candidates:
        cand_norm = re.sub(r"\s+", "", cand)
        if cand_norm in header:
            idx_on = header.index(cand_norm)
            break

    rows: list[dict] = []
    for row in data_rows:
        keyword = row[idx_kw].strip() if len(row) > idx_kw else ""
        if not keyword:
            continue
        raw_on = row[idx_on].strip() if idx_on is not None and len(row) > idx_on else "Y"
        active = str(raw_on).strip().upper() in {"Y", "YES", "TRUE", "1", "\ud65c\uc131", "\uc0ac\uc6a9"}
        rows.append({NA_COL_KEYWORD: keyword, NA_COL_ACTIVE: active})

    if not rows:
        rows = [{NA_COL_KEYWORD: kw, NA_COL_ACTIVE: True} for kw in DEFAULT_NORTH_AMERICA_KEYWORDS]
    active_keywords = [r[NA_COL_KEYWORD] for r in rows if r.get(NA_COL_ACTIVE)]
    return rows, active_keywords


def main():
    run_perf_rows: list[dict] = []
    run_step_start = perf_counter()
    inject_css()
    initialize_state()
    run_step_start = record_perf_step(run_perf_rows, "initialize_state", run_step_start)
    try:
        data = load_dashboard_data(
            int(st.session_state.get("etc_amount_threshold", 100_000_000)),
            product_family=st.session_state.get("product_family", "\ucda9\uc8fc"),
            export_only_nonstock_custom=bool(st.session_state.get("export_only_nonstock_custom", True)),
        )
        run_step_start = record_perf_step(run_perf_rows, "load_dashboard_data_call", run_step_start)
    except ValueError as exc:
        st.error(f"\ub370\uc774\ud130 \uc2dc\ud2b8 \ub85c\ub4dc \uc2e4\ud328: {exc}")
        st.info(
            "\uad6c\uae00 \uc2dc\ud2b8\uc5d0 \ud544\uc218 \uc6cc\ud06c\uc2dc\ud2b8(\uc218\uc8fc\ub0b4\uc5ed\uc815\ubcf4/\uc218\uc8fc\uad00\ub9ac/\uacf5\uc815 \uc9c4\ud589\uc815\ubcf4)\uac00 \uc874\uc7ac\ud558\ub294\uc9c0 \ud655\uc778\ud574\uc8fc\uc138\uc694. "
            "\uc5c5\ub85c\ub4dc \uc9c1\ud6c4 \uc6a9\ub7c9 \uc815\ub9ac\uc5d0\uc11c \uc2dc\ud2b8\uac00 \uc0ad\uc81c\ub41c \uacbd\uc6b0, \ub2e4\uc2dc \uc5c5\ub85c\ub4dc\ud558\uba74 \ubcf5\uad6c\ub429\ub2c8\ub2e4."
        )
        st.stop()
    available_months = data["available_months"] or [today_kst().strftime("%Y-%m")]
    current_month = today_kst().strftime("%Y-%m")
    if st.session_state["selected_month"] not in available_months:
        st.session_state["selected_month"] = current_month if current_month in available_months else available_months[-1]
    if st.session_state["selected_order_id"] not in {order["id"] for order in data["orders"]}:
        st.session_state["selected_order_id"] = data["orders"][0]["id"] if data["orders"] else ""
    valid_ids = {order["id"] for order in data["orders"]}
    selected_order_ids = st.session_state.get("selected_order_ids", [])
    if not isinstance(selected_order_ids, list):
        selected_order_ids = []
    selected_order_ids = [order_id for order_id in selected_order_ids if order_id in valid_ids]
    if st.session_state["selected_order_id"] and st.session_state["selected_order_id"] not in selected_order_ids:
        selected_order_ids = [st.session_state["selected_order_id"]]
    if not selected_order_ids and data["orders"]:
        selected_order_ids = [data["orders"][0]["id"]]
    st.session_state["selected_order_ids"] = selected_order_ids
    if st.session_state["detail_selected_order_id"] not in {order["id"] for order in data["orders"]}:
        st.session_state["detail_selected_order_id"] = st.session_state["selected_order_id"]
    run_step_start = record_perf_step(run_perf_rows, "session_state_sync", run_step_start)

    render_header(data)
    run_step_start = record_perf_step(run_perf_rows, "render_header", run_step_start)
    filtered_orders = get_filtered_orders(data)
    run_step_start = record_perf_step(run_perf_rows, "get_filtered_orders", run_step_start)

    if not filtered_orders:
        # 필터 결과가 없어도 캘린더/상세 레이아웃은 유지해서
        # 사용자가 빈 상태를 동일한 화면 구조에서 확인할 수 있게 한다.
        st.session_state["detail_metric"] = None
        st.session_state["detail_order_ids"] = []
        st.session_state["drilldown_order_id"] = None

    render_metrics(filtered_orders)
    run_step_start = record_perf_step(run_perf_rows, "render_metrics", run_step_start)
    # Calendar interaction can update/clear popup state.
    # Render dialogs after calendar handling to avoid popup flash.
    render_calendar_and_detail(filtered_orders, data, available_months)
    run_step_start = record_perf_step(run_perf_rows, "render_calendar_and_detail", run_step_start)
    render_dialogs(filtered_orders, data)
    run_step_start = record_perf_step(run_perf_rows, "render_dialogs", run_step_start)

    if SHOW_PERF_LOG_PANEL:
        load_perf_rows = data.get("perf_load_ms", []) if isinstance(data, dict) else []
        with st.sidebar.expander("실행 성능 로그", expanded=False):
            if load_perf_rows:
                st.caption("load_dashboard_data 내부 단계 (ms)")
                st.dataframe(pd.DataFrame(load_perf_rows), use_container_width=True, hide_index=True)
            st.caption("main 렌더 단계 (ms)")
            st.dataframe(pd.DataFrame(run_perf_rows), use_container_width=True, hide_index=True)


def render_order_list(filtered_orders: list[dict]):
    with st.container(border=True):
        st.markdown("## 수주건 리스트")
        cols = st.columns(3)
        for idx, order in enumerate(filtered_orders):
            with cols[idx % 3]:
                is_selected = st.session_state.get("selected_order_id") == order["id"]
                with st.container(border=True):
                    if is_selected:
                        st.markdown(
                            "<div style='margin-bottom:10px; font-size:12px; font-weight:700; color:#0f172a;'>현재 선택된 수주건</div>",
                            unsafe_allow_html=True,
                        )
                    else:
                        st.markdown("<div style='height:26px;'></div>", unsafe_allow_html=True)

                    st.markdown(
                        style_badge(order["type"], TYPE_COLORS[order["type"]])
                        + style_badge(order["risk"], RISK_COLORS[order["risk"]])
                        + (style_badge("북미", NA_COLORS) if order.get("isNorthAmerica") else ""),
                        unsafe_allow_html=True,
                    )
                    st.markdown(f"### {order['displayName']}")
                    st.caption(order.get("secondaryName", ""))
                    st.markdown(
                        "<div class='tiny'>아래 카드 버튼으로 선택과 상세 팝업을 함께 엽니다.</div>",
                        unsafe_allow_html=True,
                    )
                    st.write(f"일정: {order['startDate']} ~ {order['endDate']}")
                    st.write(f"사업소: {order.get('office', '-')}")
                    st.write(f"통합 수주건 수: {order.get('groupedOrders', 0)}")
                    st.write(f"현장: {order.get('site', '-')}")
                    if st.button("이 수주건 보기", key=f"order_pick_{order['id']}", use_container_width=True):
                        st.session_state["selected_order_id"] = order["id"]
                        st.session_state["drilldown_order_id"] = order["id"]
                        st.rerun()


def render_metrics(filtered_orders: list[dict]):
    selected_month = st.session_state["selected_month"]
    year, month = map(int, selected_month.split("-"))
    today = today_kst()
    ref_date = today if (today.year == year and today.month == month) else date(year, month, 1)

    week_no = ((ref_date.day - 1) // 7) + 1
    week_start = ref_date - timedelta(days=ref_date.weekday())
    week_end = week_start + timedelta(days=6)

    def overlaps_range(order: dict, start_day: date, end_day: date) -> bool:
        start = date.fromisoformat(order["startDate"])
        end = date.fromisoformat(order["endDate"])
        return not (end < start_day or start > end_day)

    weekly_orders = [order for order in filtered_orders if overlaps_range(order, week_start, week_end)]
    monthly_orders = filtered_orders
    biweekly_start = date(year, month, 1)
    biweekly_end = date(year, month, 14)
    north_america_biweekly_orders = [
        order
        for order in filtered_orders
        if order.get("isNorthAmerica") and overlaps_range(order, biweekly_start, biweekly_end)
    ]

    st.session_state["weekly_order_ids"] = [order["id"] for order in weekly_orders]
    st.session_state["monthly_order_ids"] = [order["id"] for order in monthly_orders]
    st.session_state["na_biweekly_order_ids"] = [order["id"] for order in north_america_biweekly_orders]

    weekly_export = sum(1 for order in weekly_orders if order["type"] == "수출")
    weekly_domestic = sum(1 for order in weekly_orders if order["type"] == "내수")
    monthly_export = sum(1 for order in monthly_orders if order["type"] == "수출")
    monthly_domestic = sum(1 for order in monthly_orders if order["type"] == "내수")

    cards = [
        {
            "key": "weekly",
            "subtitle": f"{str(year)[-2:]}년 {month}월 {week_no}주 주요 수주건",
            "total": f"{len(weekly_orders):,}",
            "partial": f"수출 {weekly_export:,}건 / 내수 {weekly_domestic:,}건",
        },
        {
            "key": "monthly",
            "subtitle": f"{str(year)[-2:]}년 {month}월 주요 수주건",
            "total": f"{len(monthly_orders):,}",
            "partial": f"수출 {monthly_export:,}건 / 내수 {monthly_domestic:,}건",
        },
        {
            "key": "na_biweekly",
            "subtitle": f"북미 수주건 {month}월 1~2주",
            "total": f"{len(north_america_biweekly_orders):,}",
            "partial": "기준: 대리점/실적대리점",
        },
    ]

    clicked_summary = summary_cards_component(
        data={"cards": cards},
        key="summary_cards_click",
        default=None,
    )
    clicked_key = None
    clicked_ts = None
    if isinstance(clicked_summary, dict):
        clicked_key = clicked_summary.get("key")
        clicked_ts = clicked_summary.get("ts")
    elif isinstance(clicked_summary, str):
        clicked_key = clicked_summary

    is_new_summary_click = False
    if clicked_ts is not None:
        try:
            clicked_ts_int = int(clicked_ts)
        except Exception:
            clicked_ts_int = 0
        if clicked_ts_int and clicked_ts_int != int(st.session_state.get("last_summary_click_ts", 0)):
            st.session_state["last_summary_click_ts"] = clicked_ts_int
            is_new_summary_click = True
    elif clicked_key:
        # 문자열 fallback(구버전 캐시)에서는 즉시 1회 처리
        is_new_summary_click = True

    if not is_new_summary_click:
        return

    if clicked_key == "weekly":
        st.session_state["detail_metric"] = "totalOrders"
        st.session_state["detail_order_ids"] = st.session_state.get("weekly_order_ids", [])
    elif clicked_key == "monthly":
        st.session_state["detail_metric"] = "groupedCount"
        st.session_state["detail_order_ids"] = st.session_state.get("monthly_order_ids", [])
    elif clicked_key == "na_biweekly":
        st.session_state["detail_metric"] = "totalOrders"
        st.session_state["detail_order_ids"] = st.session_state.get("na_biweekly_order_ids", [])


def get_filtered_orders(data: dict):
    month_value = st.session_state["selected_month"]
    business_type = st.session_state["business_type"]
    year, month = map(int, month_value.split("-"))
    month_start = date(year, month, 1)
    month_end = date(year, month, calendar.monthrange(year, month)[1])
    filtered = []
    for order in data["orders"]:
        try:
            order_start = date.fromisoformat(str(order.get("startDate", "")))
            order_end = date.fromisoformat(str(order.get("endDate", "")))
            if order_end < order_start:
                order_start, order_end = order_end, order_start
            month_match = not (order_end < month_start or order_start > month_end)
        except Exception:
            month_match = False
        if business_type == "전체":
            type_match = not bool(order.get("isEtc"))
        elif business_type == "내수":
            type_match = order["type"] == "내수" and not bool(order.get("isEtc"))
        elif business_type == "수출":
            # 수출은 북미 포함
            if bool(st.session_state.get("export_only_nonstock_custom", True)):
                type_match = order["type"] == "수출" and not bool(order.get("isEtc"))
            else:
                type_match = order["type"] == "수출"
        elif business_type == "북미":
            # 북미는 수출 중 북미건만
            type_match = order["type"] == "수출" and bool(order.get("isNorthAmerica")) and not bool(order.get("isEtc"))
        elif business_type == "기타":
            type_match = bool(order.get("isEtc"))
        else:
            type_match = order["type"] == business_type
        if month_match and type_match:
            filtered.append(order)
    return filtered


def build_overlay_calendar_payload(filtered_orders: list[dict], selected_month: str, view_style: str):
    year, month = map(int, selected_month.split("-"))
    month_calendar = calendar.Calendar(firstweekday=6).monthdatescalendar(year, month)
    today = today_kst()
    weeks = []

    for week_days in month_calendar:
        week_start = week_days[0]
        week_end = week_days[-1]
        week_events = []

        for order in filtered_orders:
            order_start = date.fromisoformat(order["startDate"])
            order_end = date.fromisoformat(order["endDate"])
            if order_end < week_start or order_start > week_end:
                continue

            visible_start = max(order_start, week_start)
            visible_end = min(order_end, week_end)
            start_col = max(0, (visible_start - week_start).days)
            end_col = min(6, (visible_end - week_start).days)

            if visible_start == order_start and visible_end == order_end:
                shape = "single"
            elif visible_start == order_start:
                shape = "start"
            elif visible_end == order_end:
                shape = "end"
            else:
                shape = "middle"

            label = order["displayName"] if view_style == "구성요소" else order["title"]
            if order.get("isNorthAmerica"):
                color_type = "north_america"
            elif order["type"] == "내수":
                color_type = "domestic"
            else:
                color_type = "export"

            week_events.append(
                {
                    "group_key": order["id"],
                    "label": clip_text(label, 22),
                    "title": f"{order['displayName']} | {order['startDate']} ~ {order['endDate']}",
                    "start_col": start_col,
                    "end_col": end_col,
                    "color_type": color_type,
                    "shape": shape,
                }
            )

        week_events.sort(key=lambda item: (item["start_col"], item["end_col"], item["label"]))
        lanes: list[list[dict]] = []
        for event in week_events:
            placed = False
            for lane in lanes:
                if all(event["end_col"] < existing["start_col"] or event["start_col"] > existing["end_col"] for existing in lane):
                    lane.append(event)
                    placed = True
                    break
            if not placed:
                lanes.append([event])

        weeks.append(
            {
                "week_key": f"{week_days[0].isoformat()}_{week_days[-1].isoformat()}",
                "days": [
                    {
                        "day": day.day,
                        "in_month": day.month == month,
                        "is_today": day == today,
                        "weekday_index": idx,
                    }
                    for idx, day in enumerate(week_days)
                ],
                "lanes": lanes,
            }
        )

    return {"weekday_labels": WEEKDAY_LABELS, "weeks": weeks}


def render_calendar_and_detail(filtered_orders: list[dict], data: dict, available_months: list[str]):
    order_by_id = {order["id"]: order for order in filtered_orders}
    valid_ids = set(order_by_id.keys())
    selected_ids = st.session_state.get("selected_order_ids", [])
    if not isinstance(selected_ids, list):
        selected_ids = []
    selected_ids = [order_id for order_id in selected_ids if order_id in valid_ids]

    selected_id = st.session_state.get("selected_order_id", "")
    if selected_id in valid_ids and selected_id not in selected_ids:
        selected_ids = [selected_id]
    if not selected_ids and filtered_orders:
        selected_ids = [filtered_orders[0]["id"]]

    st.session_state["selected_order_ids"] = selected_ids
    if selected_ids and st.session_state.get("selected_order_id") not in selected_ids:
        st.session_state["selected_order_id"] = selected_ids[0]
    if st.session_state.get("detail_selected_order_id") != st.session_state.get("selected_order_id"):
        st.session_state["detail_selected_order_id"] = st.session_state.get("selected_order_id")

    left_col, right_col = st.columns([1, 1])
    with left_col:
        with st.container(border=True):
            st.radio(
                "\uc0ac\uc5c5\uc7a5",
                options=["충주", "안성"],
                horizontal=True,
                key="product_family",
                on_change=on_top_filter_change,
            )
            filter_col1, filter_col2 = st.columns([1.35, 1])
            with filter_col1:
                st.radio(
                    "구분",
                    options=["전체", "내수", "수출", "북미", "기타"],
                    horizontal=True,
                    key="business_type",
                    on_change=on_top_filter_change,
                )
            with filter_col2:
                st.selectbox(
                    "조회 월",
                    options=available_months,
                    format_func=month_label,
                    key="selected_month",
                    on_change=on_top_filter_change,
                )
            if st.session_state.get("business_type") == "수출":
                mode_col, _ = st.columns([1.7, 1])
                with mode_col:
                    if hasattr(st, "segmented_control"):
                        export_mode = st.segmented_control(
                            "수출 기준",
                            options=["전체", "비재고/주문품"],
                            key="export_filter_mode",
                        )
                    else:
                        export_mode = st.radio(
                            "수출 기준",
                            options=["전체", "비재고/주문품"],
                            horizontal=True,
                            key="export_filter_mode",
                        )
                export_strict_on = export_mode == "비재고/주문품"
                if export_strict_on != bool(st.session_state.get("export_only_nonstock_custom", False)):
                    st.session_state["export_only_nonstock_custom"] = export_strict_on
                    st.rerun()
            if st.session_state.get("business_type") == "기타":
                current_threshold = int(st.session_state.get("etc_amount_threshold", 100_000_000))
                if "etc_amount_threshold_draft" not in st.session_state:
                    st.session_state["etc_amount_threshold_draft"] = current_threshold

                with st.form("etc_amount_threshold_form", clear_on_submit=False):
                    st.number_input(
                        "기타 금액 기준(원)",
                        min_value=0,
                        step=10_000_000,
                        key="etc_amount_threshold_draft",
                        help="충주1/충주2/F우레탄 제품의 통합 수주금액 합계 기준입니다.",
                    )
                    apply_threshold = st.form_submit_button("기준 적용", use_container_width=True)

                draft_threshold = int(st.session_state.get("etc_amount_threshold_draft", current_threshold))
                st.caption(
                    f"설정값: {draft_threshold:,}원 ({format_korean_amount_unit(draft_threshold)}원)"
                )

                if apply_threshold and draft_threshold != current_threshold:
                    st.session_state["etc_amount_threshold"] = draft_threshold
                    st.rerun()
            payload = build_overlay_calendar_payload(
                filtered_orders,
                st.session_state["selected_month"],
                st.session_state["view_style"],
            )
            clicked_result = overlay_calendar_component(
                data=payload,
                selected_group_key=st.session_state["selected_order_id"],
                selected_group_keys=st.session_state.get("selected_order_ids", []),
                key="overlay_calendar",
                default=None,
            )
            clicked_order_id = None
            selected_keys_from_component = None
            clicked_ts_int = 0
            if isinstance(clicked_result, dict):
                clicked_order_id = clicked_result.get("lastClicked")
                raw_keys = clicked_result.get("selectedKeys")
                if isinstance(raw_keys, list):
                    selected_keys_from_component = [
                        str(order_id) for order_id in raw_keys if str(order_id) in valid_ids
                    ]
                try:
                    clicked_ts_int = int(clicked_result.get("ts") or 0)
                except Exception:
                    clicked_ts_int = 0
            elif isinstance(clicked_result, str):
                clicked_order_id = clicked_result
                if clicked_order_id in valid_ids:
                    selected_keys_from_component = [clicked_order_id]

            last_calendar_ts = int(st.session_state.get("last_calendar_click_ts", 0))
            is_new_calendar_click = clicked_ts_int > 0 and clicked_ts_int != last_calendar_ts
            if is_new_calendar_click:
                st.session_state["last_calendar_click_ts"] = clicked_ts_int
            calendar_bar_clicked = (
                is_new_calendar_click
                and bool(clicked_order_id)
                and clicked_order_id in valid_ids
            )

            changed = False
            if calendar_bar_clicked and selected_keys_from_component is not None:
                if selected_keys_from_component != st.session_state.get("selected_order_ids", []):
                    st.session_state["selected_order_ids"] = selected_keys_from_component
                    changed = True
            if calendar_bar_clicked and clicked_order_id != st.session_state["selected_order_id"]:
                st.session_state["selected_order_id"] = clicked_order_id
                st.session_state["detail_selected_order_id"] = clicked_order_id
                if clicked_order_id not in st.session_state.get("selected_order_ids", []):
                    st.session_state["selected_order_ids"] = [clicked_order_id]
                changed = True
            if calendar_bar_clicked and st.session_state.get("detail_metric") is not None:
                opened_ts = int(st.session_state.get("detail_metric_open_ts", 0))
                # Close summary popup only on a genuinely newer calendar click.
                if clicked_ts_int > opened_ts:
                    st.session_state["detail_metric"] = None
                    st.session_state["detail_order_ids"] = []
                    changed = True
            if changed:
                st.rerun()

    with right_col:
        with st.container(border=True):
            selected_order = next((order for order in filtered_orders if order["id"] == st.session_state["selected_order_id"]), None)
            if not selected_order:
                st.info("선택 가능한 수주건이 없습니다.")
                return

            selected_orders = [
                order_by_id[order_id]
                for order_id in st.session_state.get("selected_order_ids", [])
                if order_id in order_by_id
            ]
            title_text = selected_order["displayName"]
            if len(selected_orders) > 1:
                title_text = f"{selected_order['displayName']} 외 {len(selected_orders) - 1}건"
            st.markdown(f"### {title_text}")
            related_rows = data["related_by_order"].get(selected_order["id"], [])
            related_options = []
            for row in related_rows:
                order_no = row.get("관련 수주번호", "")
                order_name = row.get("관련 수주건명", "")
                if not order_no and not order_name:
                    continue
                label = order_name if order_name else order_no
                if order_no and order_name:
                    label = f"{order_name} ({order_no})"
                related_options.append({"label": label, "order_no": order_no})
            option_labels = [opt["label"] for opt in related_options]
            with st.expander("세부 수주건 필터", expanded=False):
                st.caption("표시할 세부 수주건을 체크/해제하세요.")
                control_cols = st.columns(2)
                if control_cols[0].button("전체 선택", key=f"related_order_select_all_{selected_order['id']}", use_container_width=True):
                    for idx in range(len(option_labels)):
                        st.session_state[f"related_order_check_{selected_order['id']}_{idx}"] = True
                    st.rerun()
                if control_cols[1].button("전체 해제", key=f"related_order_clear_all_{selected_order['id']}", use_container_width=True):
                    for idx in range(len(option_labels)):
                        st.session_state[f"related_order_check_{selected_order['id']}_{idx}"] = False
                    st.rerun()
                selected_related_labels = []
                with st.container(height=220):
                    for idx, label in enumerate(option_labels):
                        checked = st.checkbox(
                            label,
                            value=True,
                            key=f"related_order_check_{selected_order['id']}_{idx}",
                        )
                        if checked:
                            selected_related_labels.append(label)
            selected_related_nos = {
                opt["order_no"]
                for opt in related_options
                if opt["label"] in selected_related_labels
            }

            badge_line = style_badge(selected_order["type"], TYPE_COLORS[selected_order["type"]])
            if selected_order.get("isNorthAmerica"):
                badge_line += style_badge("북미", NA_COLORS)
            st.markdown(badge_line, unsafe_allow_html=True)

            summary_source_orders = selected_orders if selected_orders else [selected_order]
            compact_cols = st.columns(3)
            due_card_placeholder = compact_cols[0].empty()
            item_card_placeholder = compact_cols[1].empty()
            qty_card_placeholder = compact_cols[2].empty()

            current_order_id = selected_order["id"]
            selected_order_rows_for_default = data.get("detail_items_by_order", {}).get(selected_order["id"], [])
            if selected_related_labels:
                selected_order_rows_for_default = [
                    row for row in selected_order_rows_for_default
                    if row.get("관련 수주번호", "") in selected_related_nos
                ]

            chgju_products = {"충주1제품", "충주2제품", "충주상품", "충주2상품", "F우레탄제품"}
            wood_products = {"F우레탄제품", "베트남상품", "목제상품", "목제5상품", "목제6상품"}
            product_family = st.session_state.get("product_family", "충주")
            export_strict_on = bool(st.session_state.get("export_only_nonstock_custom", False))
            standard_options = ["전체", "주문품"] if export_strict_on else ["전체", "표준품", "주문품"]
            product_options = ["전체", "안성제품", "안성상품"] if product_family == "안성" else ["전체", "충주", "목제상품"]

            if st.session_state.get("item_filter_last_order_id") != current_order_id:
                if selected_order.get("isNorthAmerica"):
                    st.session_state[f"item_standard_filter_{current_order_id}"] = "전체"
                    st.session_state[f"item_product_filter_{current_order_id}"] = "전체"
                elif selected_order.get("type") == "내수":
                    st.session_state[f"item_standard_filter_{current_order_id}"] = "주문품" if export_strict_on else "전체"
                    st.session_state[f"item_product_filter_{current_order_id}"] = "충주" if product_family != "안성" else "안성제품"
                else:
                    st.session_state[f"item_standard_filter_{current_order_id}"] = "주문품" if export_strict_on else "전체"
                    default_product = "전체"
                    if selected_order_rows_for_default:
                        default_df = pd.DataFrame(selected_order_rows_for_default)
                        if export_strict_on and "표준구분" in default_df.columns:
                            default_df = default_df[
                                default_df["표준구분"].astype(str).str.contains("주문품", na=False)
                            ]
                        product_values = set(default_df.get("제품구분", pd.Series(dtype=str)).astype(str))
                        if product_family == "안성":
                            if any("안성제품" in str(product) for product in product_values):
                                default_product = "안성제품"
                            elif any("안성상품" in str(product) for product in product_values):
                                default_product = "안성상품"
                        else:
                            if any(product in chgju_products for product in product_values):
                                default_product = "충주"
                            elif any(product in wood_products for product in product_values):
                                default_product = "목제상품"
                    st.session_state[f"item_product_filter_{current_order_id}"] = default_product
                st.session_state["item_filter_last_order_id"] = current_order_id

            current_standard_key = f"item_standard_filter_{current_order_id}"
            current_product_key = f"item_product_filter_{current_order_id}"
            if st.session_state.get(current_product_key) == "목제":
                st.session_state[current_product_key] = "목제상품"
            if st.session_state.get(current_standard_key) not in standard_options:
                st.session_state[current_standard_key] = "주문품" if export_strict_on else "전체"
            if st.session_state.get(current_product_key) not in product_options:
                st.session_state[current_product_key] = product_options[0]

            filter_cols = st.columns(3)
            standard_filter = filter_cols[0].selectbox(
                "표준구분",
                options=standard_options,
                key=current_standard_key,
            )
            product_filter = filter_cols[1].selectbox(
                "제품구분",
                options=product_options,
                key=current_product_key,
            )
            return_only_filter = filter_cols[2].toggle(
                "회수",
                key=f"item_return_filter_{current_order_id}",
            )

            def build_item_display_df(
                raw_rows: list[dict],
                selected_standard_filter: str | None = None,
                selected_product_filter: str | None = None,
                selected_return_only_filter: bool | None = None,
            ) -> pd.DataFrame:
                df = pd.DataFrame(raw_rows)
                if df.empty:
                    return df
                effective_standard = standard_filter if selected_standard_filter is None else selected_standard_filter
                effective_product = product_filter if selected_product_filter is None else selected_product_filter
                effective_return_only = return_only_filter if selected_return_only_filter is None else selected_return_only_filter

                if effective_standard in {"주문품", "표준품"} and "표준구분" in df.columns:
                    df = df[df["표준구분"].astype(str).str.contains(effective_standard, na=False)]
                if "제품구분" in df.columns:
                    if product_family == "안성":
                        family_mask = df["제품구분"].astype(str).str.contains("안성", na=False)
                    else:
                        # 충주 화면의 "전체"는 기존 "목제상품" 필터(충주+목제군)를 의미.
                        family_mask = df["제품구분"].isin(chgju_products | wood_products)

                    if effective_product == "전체":
                        df = df[family_mask]
                    elif effective_product == "충주":
                        df = df[df["제품구분"].isin(chgju_products)]
                    elif effective_product == "목제상품":
                        df = df[df["제품구분"].isin(wood_products)]
                    elif effective_product == "안성제품":
                        df = df[
                            df["제품구분"].astype(str).str.contains("안성", na=False)
                            & df["제품구분"].astype(str).str.contains("제품", na=False)
                        ]
                    elif effective_product == "안성상품":
                        df = df[
                            df["제품구분"].astype(str).str.contains("안성", na=False)
                            & df["제품구분"].astype(str).str.contains("상품", na=False)
                        ]
                df = df.rename(columns={"품목명": "단품명칭", "수량": "수주량"})
                if effective_return_only and "현재고" in df.columns and "수주량" in df.columns:
                    order_qty_series = to_numeric(df["수주량"])
                    if "예량(D+7)" in df.columns:
                        d7_min_forecast = to_numeric(df["예량(D+7)"])
                        # 회수 기준: 달력일 D+7 이내의 최소 예량 - 수주량 >= 0
                        df = df[(d7_min_forecast - order_qty_series) >= 0]
                    else:
                        stock_series = to_numeric(df["현재고"])
                        # 예량 데이터가 없으면 기존 현재고 기준으로 fallback.
                        df = (df[(stock_series - order_qty_series) >= 0])
                if "단품명칭" in df.columns:
                    df["단품명칭"] = df["단품명칭"].map(shorten_item_name_for_display)
                wanted_cols = [
                    "대표 수주건명",
                    "브랜드",
                    "제품구분",
                    "단품코드",
                    "색상",
                    "단품명칭",
                    "수주량",
                    "현재고",
                    "확정납기",
                    "예량(D+7)",
                    "출고예정(D+7)",
                ]
                return df[[col for col in wanted_cols if col in df.columns]]

            merged_item_rows = []
            for each_order in summary_source_orders:
                each_rows = data.get("detail_items_by_order", {}).get(each_order["id"], [])
                if each_order["id"] == selected_order["id"] and selected_related_labels:
                    each_rows = [
                        row for row in each_rows
                        if row.get("관련 수주번호", "") in selected_related_nos
                    ]
                for row in each_rows:
                    row_copy = dict(row)
                    row_copy["대표 수주건명"] = each_order.get("displayName", "")
                    merged_item_rows.append(row_copy)
            display_df = build_item_display_df(merged_item_rows)
            if display_df.empty and merged_item_rows and product_filter != "전체":
                fallback_df = build_item_display_df(
                    merged_item_rows,
                    selected_standard_filter=standard_filter,
                    selected_product_filter="전체",
                    selected_return_only_filter=return_only_filter,
                )
                if not fallback_df.empty:
                    st.info("현재 제품구분 필터 결과가 없어, 제품구분 전체 기준 품목을 표시합니다.")
                    display_df = fallback_df

            # Summary cards should follow exactly the same filtered rows shown in the table.
            if not display_df.empty and "확정납기" in display_df.columns:
                due_series = pd.to_datetime(display_df["확정납기"], errors="coerce").dropna()
            else:
                due_series = pd.Series(dtype="datetime64[ns]")

            if not due_series.empty:
                min_start = due_series.min().date().isoformat()
                max_end = due_series.max().date().isoformat()
            else:
                min_start = min(order.get("startDate", selected_order["startDate"]) for order in summary_source_orders)
                max_end = max(order.get("endDate", selected_order["endDate"]) for order in summary_source_orders)

            if not display_df.empty:
                if {"단품코드", "색상"}.issubset(display_df.columns):
                    total_items = int(
                        display_df[["단품코드", "색상"]]
                        .astype(str)
                        .drop_duplicates()
                        .shape[0]
                    )
                else:
                    total_items = int(len(display_df))
                if "수주량" in display_df.columns:
                    total_amount = int(pd.to_numeric(display_df["수주량"], errors="coerce").fillna(0).sum())
                else:
                    total_amount = 0
            else:
                total_items = 0
                total_amount = 0

            due_card_placeholder.markdown(
                f"""
                <div class="soft-card" style="padding:10px 12px; min-height:unset;">
                    <div class="subtle-title" style="font-size:12px;">확정납기</div>
                    <div style="font-size:13px; font-weight:700; color:#0f172a; line-height:1.35;">
                        {format_korean_date(min_start)}<br>
                        ~{format_korean_date(max_end)}
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )
            item_card_placeholder.markdown(
                f"""
                <div class="soft-card" style="padding:10px 12px; min-height:unset;">
                    <div class="subtle-title" style="font-size:12px;">합계 품목수</div>
                    <div style="font-size:13px; font-weight:700; color:#0f172a;">
                        {total_items:,} 품목
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )
            qty_card_placeholder.markdown(
                f"""
                <div class="soft-card" style="padding:10px 12px; min-height:unset;">
                    <div class="subtle-title" style="font-size:12px;">합계 수주량</div>
                    <div style="font-size:13px; font-weight:700; color:#0f172a;">
                        {total_amount:,} BOX
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )
            display_df_view = with_row_no_and_total(display_df)
            st.dataframe(format_df_numbers_for_display(display_df_view), use_container_width=True, hide_index=True)
            if not display_df.empty:
                excel_bytes = dataframe_to_styled_excel_bytes(display_df_view, sheet_name="통합품목")
                st.download_button(
                    "다운로드 (.xlsx)",
                    data=excel_bytes,
                    file_name=f"통합품목_{today_kst().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=False,
                )

def render_metrics(all_orders: list[dict]):
    today = today_kst()
    year, month = today.year, today.month

    month_weeks = calendar.Calendar(firstweekday=0).monthdatescalendar(year, month)
    week_no = 1
    week_start, week_end = today, today
    for idx, week in enumerate(month_weeks, start=1):
        if today in week:
            week_no = idx
            week_start, week_end = week[0], week[-1]
            break

    month_start = date(year, month, 1)
    month_end = date(year, month, calendar.monthrange(year, month)[1])
    biweekly_start = today
    biweekly_end = today + timedelta(days=14)

    def overlaps_range(order: dict, start_day: date, end_day: date) -> bool:
        start = date.fromisoformat(order["startDate"])
        end = date.fromisoformat(order["endDate"])
        return not (end < start_day or start > end_day)

    weekly_orders = [order for order in all_orders if overlaps_range(order, week_start, week_end)]
    monthly_orders = [order for order in all_orders if overlaps_range(order, month_start, month_end)]
    north_america_biweekly_orders = [
        order
        for order in all_orders
        if order.get("isNorthAmerica") and overlaps_range(order, biweekly_start, biweekly_end)
    ]

    st.session_state["weekly_order_ids"] = [order["id"] for order in weekly_orders]
    st.session_state["monthly_order_ids"] = [order["id"] for order in monthly_orders]
    st.session_state["na_biweekly_order_ids"] = [order["id"] for order in north_america_biweekly_orders]

    weekly_export = sum(1 for order in weekly_orders if order["type"] == "수출")
    weekly_domestic = sum(1 for order in weekly_orders if order["type"] == "내수")
    monthly_export = sum(1 for order in monthly_orders if order["type"] == "수출")
    monthly_domestic = sum(1 for order in monthly_orders if order["type"] == "내수")

    cards = [
        {
            "key": "weekly",
            "subtitle": f"{str(year)[-2:]}년 {month}월 {week_no}주 주요 수주건",
            "total": f"{len(weekly_orders):,}",
            "partial": f"수출 {weekly_export:,}건 / 내수 {weekly_domestic:,}건",
        },
        {
            "key": "monthly",
            "subtitle": f"{str(year)[-2:]}년 {month}월 주요 수주건",
            "total": f"{len(monthly_orders):,}",
            "partial": f"수출 {monthly_export:,}건 / 내수 {monthly_domestic:,}건",
        },
        {
            "key": "na_biweekly",
            "subtitle": f"북미 수주건 {month}월 1~2주",
            "total": f"{len(north_america_biweekly_orders):,}",
            "partial": "기준: 대리점/실적대리점",
        },
    ]

    clicked_summary = summary_cards_component(
        data={"cards": cards},
        key="summary_cards_click",
        default=None,
    )
    if clicked_summary == "weekly":
        st.session_state["detail_metric"] = "totalOrders"
        st.session_state["detail_order_ids"] = st.session_state.get("weekly_order_ids", [])
    elif clicked_summary == "monthly":
        st.session_state["detail_metric"] = "groupedCount"
        st.session_state["detail_order_ids"] = st.session_state.get("monthly_order_ids", [])
    elif clicked_summary == "na_biweekly":
        st.session_state["detail_metric"] = "totalOrders"
        st.session_state["detail_order_ids"] = st.session_state.get("na_biweekly_order_ids", [])

def render_metrics(all_orders: list[dict]):
    today = today_kst()
    year, month = today.year, today.month

    month_weeks = calendar.Calendar(firstweekday=0).monthdatescalendar(year, month)
    week_no = 1
    week_start, week_end = today, today
    for idx, week in enumerate(month_weeks, start=1):
        if today in week:
            week_no = idx
            week_start, week_end = week[0], week[-1]
            break

    month_start = date(year, month, 1)
    month_end = date(year, month, calendar.monthrange(year, month)[1])
    due_soon_start = today
    due_soon_end = today + timedelta(days=3)

    def overlaps_range(order: dict, start_day: date, end_day: date) -> bool:
        start = date.fromisoformat(order["startDate"])
        end = date.fromisoformat(order["endDate"])
        return not (end < start_day or start > end_day)

    weekly_orders = [order for order in all_orders if overlaps_range(order, week_start, week_end)]
    monthly_orders = [order for order in all_orders if overlaps_range(order, month_start, month_end)]
    due_soon_orders = []
    for order in all_orders:
        try:
            due_day = date.fromisoformat(order["endDate"])
        except Exception:
            continue
        if due_soon_start <= due_day <= due_soon_end:
            due_soon_orders.append(order)

    st.session_state["weekly_order_ids"] = [order["id"] for order in weekly_orders]
    st.session_state["monthly_order_ids"] = [order["id"] for order in monthly_orders]
    st.session_state["due_soon_order_ids"] = [order["id"] for order in due_soon_orders]

    weekly_export = sum(1 for order in weekly_orders if order["type"] == "수출")
    weekly_domestic = sum(1 for order in weekly_orders if order["type"] == "내수")
    monthly_export = sum(1 for order in monthly_orders if order["type"] == "수출")
    monthly_domestic = sum(1 for order in monthly_orders if order["type"] == "내수")

    cards = [
        {
            "key": "weekly",
            "subtitle": f"{str(year)[-2:]}년 {month}월 {week_no}주 주요 수주건",
            "total": f"{len(weekly_orders):,}",
            "partial": f"수출 {weekly_export:,}건 / 내수 {weekly_domestic:,}건",
        },
        {
            "key": "monthly",
            "subtitle": f"{str(year)[-2:]}년 {month}월 주요 수주건",
            "total": f"{len(monthly_orders):,}",
            "partial": f"수출 {monthly_export:,}건 / 내수 {monthly_domestic:,}건",
        },
        {
            "key": "due_soon",
            "subtitle": "납기 임박건 (D-3)",
            "total": f"{len(due_soon_orders):,}",
            "partial": f"기준: {due_soon_start.month}/{due_soon_start.day}~{due_soon_end.month}/{due_soon_end.day}",
        },
    ]

    clicked_summary = summary_cards_component(
        data={"cards": cards},
        key="summary_cards_click",
        default=None,
    )
    clicked_summary_key = None
    clicked_summary_ts = 0
    if isinstance(clicked_summary, dict):
        clicked_summary_key = clicked_summary.get("key")
        try:
            clicked_summary_ts = int(clicked_summary.get("ts") or 0)
        except Exception:
            clicked_summary_ts = 0
    elif isinstance(clicked_summary, str):
        clicked_summary_key = clicked_summary

    last_summary_ts = int(st.session_state.get("last_summary_click_ts", 0))
    is_new_summary_click = clicked_summary_ts > 0 and clicked_summary_ts != last_summary_ts
    if is_new_summary_click:
        st.session_state["last_summary_click_ts"] = clicked_summary_ts

    if not is_new_summary_click:
        return

    if clicked_summary_key == "weekly":
        st.session_state["detail_metric"] = "totalOrders"
        st.session_state["detail_order_ids"] = st.session_state.get("weekly_order_ids", [])
        st.session_state["detail_metric_open_ts"] = clicked_summary_ts
    elif clicked_summary_key == "monthly":
        st.session_state["detail_metric"] = "groupedCount"
        st.session_state["detail_order_ids"] = st.session_state.get("monthly_order_ids", [])
        st.session_state["detail_metric_open_ts"] = clicked_summary_ts
    elif clicked_summary_key == "due_soon":
        st.session_state["detail_metric"] = "totalOrders"
        st.session_state["detail_order_ids"] = st.session_state.get("due_soon_order_ids", [])
        st.session_state["detail_metric_open_ts"] = clicked_summary_ts


if __name__ == "__main__":
    main()
